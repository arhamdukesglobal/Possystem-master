# Employee.py
from tkinter import *
from PIL import Image, ImageTk, Image as PILImage
from tkinter import ttk, messagebox, filedialog
import sqlite3
from datetime import datetime
import re
from tkcalendar import Calendar  # replaced DateEntry with Calendar
import os
import locale
import tempfile
import shutil
from fpdf import FPDF  # pip install fpdf
import pandas as pd  # pip install pandas openpyxl
import base64
import io
import time


class EmployeeClass:
    def __init__(self, root, dashboard_window=None):
        self.root = root
        self.dashboard_window = dashboard_window  # Store reference to dashboard

        # Center the window on screen and set appropriate size for laptop
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        window_width = 1100
        window_height = 700  # Slightly reduced height
        x_position = (screen_width - window_width) // 2
        y_position = (screen_height - window_height) // 2

        self.root.geometry(
            f"{window_width}x{window_height}+{x_position}+{y_position}")
        self.root.title(
            "Inventory Management System | Developed by Dukes Tech Services")
        self.root.config(bg="#F5F7FA")
        self.root.focus_force()

        # Handle window closing
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

        # Variables for image paths
        self.id_card_path = ""
        self.utility_bill_path = ""

        # Variables for image data (for database storage)
        self.id_card_data = None
        self.utility_bill_data = None

        # Initialize database
        self.initialize_database()

        # ===========================================
        # All Variables=========
        self.var_searchby = StringVar()
        self.var_searchtxt = StringVar()

        self.var_EmpID = StringVar()
        self.var_gender = StringVar()
        self.var_contact = StringVar()
        self.var_name = StringVar()
        self.var_DOB = StringVar()
        self.var_DOJ = StringVar()
        self.var_email = StringVar()
        self.var_salary = StringVar()
        self.var_cnic = StringVar()

        # Main container for centering
        main_container = Frame(self.root, bg="#F5F7FA")
        main_container.pack(fill=BOTH, expand=True, padx=15, pady=10)

        # === Real-time Date & Time Frame ===
        datetime_frame = Frame(main_container, bg="#2ECC71", height=35)
        datetime_frame.pack(fill=X, padx=5, pady=(0, 5))
        datetime_frame.pack_propagate(False)

        self.datetime_label = Label(datetime_frame,
                                     font=("Segoe UI", 11, "bold"),
                                     bg="#2ECC71", fg="white")
        self.datetime_label.pack(expand=True, fill=BOTH)

        # Update date and time
        self.update_datetime()

        # ===searchFrame=====
        SearchFrame = LabelFrame(main_container, text="Search Employee",
                                 font=("Segoe UI", 10, "bold"), bd=2,
                                 relief=RIDGE, bg="white", fg="#000000")
        SearchFrame.pack(fill=X, padx=5, pady=(0, 5))

        # Center the search frame content
        search_content = Frame(SearchFrame, bg="white")
        search_content.pack(pady=8)

        # ===options====
        cmb_search = ttk.Combobox(search_content, textvariable=self.var_searchby,
                                  values=("Select", "Emp-ID", "Name",
                                          "Email", "Contact", "CNIC"),
                                  state="readonly", justify=CENTER,
                                  font=("Segoe UI", 10), width=12)
        cmb_search.pack(side=LEFT, padx=(0, 8))
        cmb_search.current(0)

        txt_search = Entry(search_content, textvariable=self.var_searchtxt,
                           font=("Segoe UI", 10), bg="#F8F9FA",
                           fg="#333333", highlightthickness=1,
                           highlightbackground="#2ECC71",
                           highlightcolor="#2ECC71", width=20)
        txt_search.pack(side=LEFT, padx=(0, 8))

        btn_search = Button(search_content, text="Search", command=self.search,
                            font=("Segoe UI", 10, "bold"), bg="#2ECC71",
                            fg="white", cursor="hand2", activebackground="#2ECC71",
                            activeforeground="white", width=12)
        btn_search.pack(side=LEFT)

        # ===title====
        title = Label(main_container, text="Employee Details",
                      font=("Segoe UI", 14, "bold"), bg="#2ECC71",
                      fg="white", padx=15, pady=8)
        title.pack(fill=X, padx=5, pady=(0, 5))

        # ===content Frame====
        content_frame = Frame(main_container, bg="#F5F7FA")
        content_frame.pack(fill=BOTH, expand=True, padx=5)

        # Left side form - fixed width to ensure table visibility
        form_frame = Frame(content_frame, bg="#F5F7FA", width=520)  # Fixed width
        form_frame.pack(side=LEFT, fill=Y, padx=(0, 10))
        form_frame.pack_propagate(False)  # Prevent expansion

        # Use grid layout for the form for perfect alignment
        # Configure grid columns with reduced widths for laptop
        form_frame.columnconfigure(1, weight=1, uniform='entry')
        form_frame.columnconfigure(3, weight=1, uniform='entry')
        form_frame.columnconfigure(0, minsize=100)  # label column 1
        form_frame.columnconfigure(2, minsize=100)  # label column 2
        # set uniform width for entry columns (slightly reduced)
        form_frame.columnconfigure(1, minsize=160)
        form_frame.columnconfigure(3, minsize=160)

        # ====row1====
        lbl_empid = Label(form_frame, text="Emp-ID",
                          font=("Segoe UI", 10, "bold"),
                          bg="#F5F7FA", fg="#000000", anchor="w")
        lbl_empid.grid(row=0, column=0, padx=(0, 5), pady=4, sticky='w')

        self.txt_empid = Entry(form_frame, textvariable=self.var_EmpID,
                               font=("Segoe UI", 10), bg="white",
                               fg="#333333", state='readonly',
                               highlightthickness=1,
                               highlightbackground="#E0E6ED",
                               highlightcolor="#2ECC71")
        self.txt_empid.grid(row=0, column=1, padx=(0, 10), pady=4, sticky='ew')

        lbl_gender = Label(form_frame, text="Gender",
                           font=("Segoe UI", 10, "bold"),
                           bg="#F5F7FA", fg="#000000", anchor="w")
        lbl_gender.grid(row=0, column=2, padx=(0, 5), pady=4, sticky='w')

        cmb_gender = ttk.Combobox(form_frame, textvariable=self.var_gender,
                                  values=("Select", "Male", "Female", "Other"),
                                  state="readonly", justify=CENTER,
                                  font=("Segoe UI", 10))
        cmb_gender.grid(row=0, column=3, padx=(0, 0), pady=4, sticky='ew')
        cmb_gender.current(0)

        # ====row2=====
        lbl_name = Label(form_frame, text="Name",
                         font=("Segoe UI", 10, "bold"),
                         bg="#F5F7FA", fg="#000000", anchor="w")
        lbl_name.grid(row=1, column=0, padx=(0, 5), pady=4, sticky='w')

        self.txt_name = Entry(form_frame, textvariable=self.var_name,
                              font=("Segoe UI", 10), bg="white",
                              fg="#333333",
                              highlightthickness=1,
                              highlightbackground="#E0E6ED",
                              highlightcolor="#2ECC71")
        self.txt_name.grid(row=1, column=1, padx=(0, 10), pady=4, sticky='ew')

        lbl_cnic = Label(form_frame, text="CNIC",
                         font=("Segoe UI", 10, "bold"),
                         bg="#F5F7FA", fg="#000000", anchor="w")
        lbl_cnic.grid(row=1, column=2, padx=(0, 5), pady=4, sticky='w')

        self.cnic_entry = Entry(form_frame, textvariable=self.var_cnic,
                                font=("Segoe UI", 10), bg="white",
                                fg="#333333",
                                highlightthickness=1,
                                highlightbackground="#E0E6ED",
                                highlightcolor="#2ECC71")
        self.cnic_entry.grid(row=1, column=3, padx=(0, 0), pady=4, sticky='ew')
        self.cnic_entry.bind('<KeyRelease>', self.validate_cnic)

        # ====row3=====  (DOB with custom date entry)
        lbl_dob = Label(form_frame, text="D.O.B",
                        font=("Segoe UI", 10, "bold"),
                        bg="#F5F7FA", fg="#000000", anchor="w")
        lbl_dob.grid(row=2, column=0, padx=(0, 5), pady=4, sticky='w')

        self.dob_frame, self.dob_entry = self.create_date_entry(
            form_frame, self.var_DOB, row=2, col=1, colspan=1)

        lbl_contact = Label(form_frame, text="Contact No.",
                            font=("Segoe UI", 10, "bold"),
                            bg="#F5F7FA", fg="#000000", anchor="w")
        lbl_contact.grid(row=2, column=2, padx=(0, 5), pady=4, sticky='w')

        contact_frame = Frame(form_frame, bg="white",
                              highlightthickness=1,
                              highlightbackground="#E0E6ED",
                              highlightcolor="#2ECC71")
        contact_frame.grid(row=2, column=3, padx=(0, 0), pady=4, sticky='ew')
        contact_frame.columnconfigure(1, weight=1)

        contact_label = Label(contact_frame, text="+92",
                              font=("Segoe UI", 10), bg="white",
                              fg="#666666", padx=3)
        contact_label.pack(side=LEFT)

        self.contact_entry = Entry(contact_frame, font=("Segoe UI", 10),
                                   bg="white", fg="#333333",
                                   justify=LEFT, relief=FLAT)
        self.contact_entry.pack(side=LEFT, fill=BOTH, expand=True)
        self.contact_entry.bind('<KeyRelease>', self.validate_contact)
        self.contact_entry.bind('<FocusIn>', self.on_contact_focus)

        # ====row4=====
        lbl_email = Label(form_frame, text="Email",
                          font=("Segoe UI", 10, "bold"),
                          bg="#F5F7FA", fg="#000000", anchor="w")
        lbl_email.grid(row=3, column=0, padx=(0, 5), pady=4, sticky='w')

        self.txt_email = Entry(form_frame, textvariable=self.var_email,
                               font=("Segoe UI", 10), bg="white",
                               fg="#333333",
                               highlightthickness=1,
                               highlightbackground="#E0E6ED",
                               highlightcolor="#2ECC71")
        self.txt_email.grid(row=3, column=1, padx=(0, 10), pady=4, sticky='ew')

        lbl_doj = Label(form_frame, text="Joining Date",
                        font=("Segoe UI", 10, "bold"),
                        bg="#F5F7FA", fg="#000000", anchor="w")
        lbl_doj.grid(row=3, column=2, padx=(0, 5), pady=4, sticky='w')

        self.doj_frame, self.doj_entry = self.create_date_entry(
            form_frame, self.var_DOJ, row=3, col=3, colspan=1)

        # ====row5=====
        lbl_salary = Label(form_frame, text="Salary",
                           font=("Segoe UI", 10, "bold"),
                           bg="#F5F7FA", fg="#000000", anchor="w")
        lbl_salary.grid(row=4, column=0, padx=(0, 5), pady=4, sticky='w')

        salary_frame = Frame(form_frame, bg="white",
                             highlightthickness=1,
                             highlightbackground="#E0E6ED",
                             highlightcolor="#2ECC71")
        salary_frame.grid(row=4, column=1, padx=(0, 10), pady=4, sticky='ew')
        salary_frame.columnconfigure(1, weight=1)

        salary_label = Label(salary_frame, text="Rs",
                             font=("Segoe UI", 10), bg="white",
                             fg="#666666", padx=3)
        salary_label.pack(side=LEFT)

        self.salary_entry = Entry(salary_frame, font=("Segoe UI", 10),
                                  bg="white", fg="#333333",
                                  justify=RIGHT, relief=FLAT)
        self.salary_entry.pack(side=LEFT, fill=BOTH, expand=True)
        self.salary_entry.bind('<KeyRelease>', self.format_salary)
        self.salary_entry.bind('<FocusOut>', self.format_salary_final)

        lbl_address = Label(form_frame, text="Address",
                            font=("Segoe UI", 10, "bold"),
                            bg="#F5F7FA", fg="#000000", anchor="w")
        lbl_address.grid(row=4, column=2, padx=(0, 5), pady=4, sticky='w')

        self.txt_address = Text(form_frame, font=("Segoe UI", 10),
                                bg="white", fg="#333333",
                                height=2, width=20,
                                highlightthickness=1,
                                highlightbackground="#E0E6ED",
                                highlightcolor="#2ECC71")
        self.txt_address.grid(row=4, column=3, padx=(0, 0), pady=4, sticky='ew')

        # ====KYE (Know Your Employee) Section=====
        kye_frame = LabelFrame(form_frame, text="KYE (Know Your Employee) - Compulsory",
                               font=("Segoe UI", 10, "bold"), bd=2,
                               relief=RIDGE, bg="white", fg="#000000")
        kye_frame.grid(row=5, column=0, columnspan=4, pady=(10, 8), sticky='ew')
        kye_frame.columnconfigure(0, weight=1)
        kye_frame.columnconfigure(1, weight=1)

        # ID Card Section
        id_card_frame = Frame(kye_frame, bg="white")
        id_card_frame.grid(row=0, column=0, padx=15, pady=10, sticky='w')

        lbl_id_card = Label(id_card_frame, text="ID Card Copy:",
                            font=("Segoe UI", 9, "bold"),
                            bg="white", fg="#000000")
        lbl_id_card.pack(anchor="w")

        self.lbl_id_status = Label(id_card_frame, text="No file selected",
                                   font=("Segoe UI", 8), bg="white",
                                   fg="red")
        self.lbl_id_status.pack(anchor="w", pady=(0, 3))

        id_button_frame = Frame(id_card_frame, bg="white")
        id_button_frame.pack(fill=X)

        btn_upload_id = Button(id_button_frame, text="Upload ID Card",
                               command=self.upload_id_card,
                               font=("Segoe UI", 8, "bold"),
                               bg="#2ECC71", fg="#ffffff",
                               cursor="hand2", width=12)
        btn_upload_id.pack(side=LEFT, padx=(0, 8))

        self.btn_view_id = Button(id_button_frame, text="View",
                                  command=self.view_id_card,
                                  font=("Segoe UI", 8, "bold"),
                                  bg="#000000", fg="#2ECC71",
                                  cursor="hand2", width=8,
                                  state=DISABLED)
        self.btn_view_id.pack(side=LEFT)

        # Utility Bill Section
        utility_frame = Frame(kye_frame, bg="white")
        utility_frame.grid(row=0, column=1, padx=15, pady=10, sticky='w')

        lbl_utility = Label(utility_frame, text="Utility Bill:",
                            font=("Segoe UI", 9, "bold"),
                            bg="white", fg="#000000")
        lbl_utility.pack(anchor="w")

        self.lbl_utility_status = Label(utility_frame, text="No file selected",
                                        font=("Segoe UI", 8), bg="white",
                                        fg="red")
        self.lbl_utility_status.pack(anchor="w", pady=(0, 3))

        utility_button_frame = Frame(utility_frame, bg="white")
        utility_button_frame.pack(fill=X)

        btn_upload_utility = Button(utility_button_frame, text="Upload Utility Bill",
                                    command=self.upload_utility_bill,
                                    font=("Segoe UI", 8, "bold"),
                                    bg="#2ECC71", fg="#ffffff",
                                    cursor="hand2", width=12)
        btn_upload_utility.pack(side=LEFT, padx=(0, 8))

        self.btn_view_utility = Button(utility_button_frame, text="View",
                                       command=self.view_utility_bill,
                                       font=("Segoe UI", 8, "bold"),
                                       bg="#000000", fg="#2ECC71",
                                       cursor="hand2", width=8,
                                       state=DISABLED)
        self.btn_view_utility.pack(side=LEFT)

        # Informational message
        info_label = Label(kye_frame,
                           text="Note: Both ID Card and Utility Bill must be uploaded before saving",
                           font=("Segoe UI", 8, "italic"),
                           bg="white", fg="#C53030")
        info_label.grid(row=1, column=0, columnspan=2, padx=15, pady=(0, 8))

        # ====buttons=====
        button_frame = Frame(form_frame, bg="#F5F7FA")
        button_frame.grid(row=6, column=0, columnspan=4, pady=(8, 5), sticky='ew')
        button_frame.columnconfigure(0, weight=1)
        button_frame.columnconfigure(1, weight=1)
        button_frame.columnconfigure(2, weight=1)
        button_frame.columnconfigure(3, weight=1)

        btn_add = Button(button_frame, text="Save", command=self.add,
                         font=("Segoe UI", 10, "bold"),
                         bg="#00C853", activebackground="#00B248",
                         fg="white", activeforeground="white",
                         cursor="hand2", bd=0, width=10)
        btn_add.grid(row=0, column=0, padx=3)

        btn_update = Button(button_frame, text="Edit/Update", command=self.update,
                            font=("Segoe UI", 10, "bold"),
                            bg="#2962FF", activebackground="#1E4ED8",
                            fg="white", activeforeground="white",
                            cursor="hand2", bd=0, width=12)
        btn_update.grid(row=0, column=1, padx=3)

        btn_delete = Button(button_frame, text="Delete", command=self.delete,
                            font=("Segoe UI", 10, "bold"),
                            bg="#FF1744", activebackground="#D50000",
                            fg="white", activeforeground="white",
                            cursor="hand2", bd=0, width=10)
        btn_delete.grid(row=0, column=2, padx=3)

        btn_clear = Button(button_frame, text="Clear", command=self.clear,
                           font=("Segoe UI", 10, "bold"),
                           bg="#FF9100", activebackground="#FF6D00",
                           fg="white", activeforeground="white",
                           cursor="hand2", bd=0, width=10)
        btn_clear.grid(row=0, column=3, padx=3)

        # Report Buttons
        report_frame = Frame(form_frame, bg="#F5F7FA")
        report_frame.grid(row=7, column=0, columnspan=4, pady=(5, 8), sticky='ew')
        report_frame.columnconfigure(0, weight=1)
        report_frame.columnconfigure(1, weight=1)

        btn_pdf = Button(report_frame, text="PDF Report",
                         command=self.generate_pdf_report,
                         font=("Segoe UI", 10, "bold"),
                         bg="#E53935", activebackground="#C62828",
                         fg="white", activeforeground="white",
                         cursor="hand2", bd=0, width=18)
        btn_pdf.grid(row=0, column=0, padx=(0, 8))

        btn_excel = Button(report_frame, text="Excel Report",
                           command=self.generate_excel_report,
                           font=("Segoe UI", 10, "bold"),
                           bg="#00B050", activebackground="#009140",
                           fg="white", activeforeground="white",
                           cursor="hand2", bd=0, width=18)
        btn_excel.grid(row=0, column=1)

        # ===Employee Details Table====
        table_frame = Frame(content_frame, bg="#000000")
        table_frame.pack(side=RIGHT, fill=BOTH, expand=True)

        emp_frame = Frame(table_frame, bd=2, relief=RIDGE, bg="white")
        emp_frame.pack(fill=BOTH, expand=True)

        scrolly = Scrollbar(emp_frame, orient=VERTICAL)
        scrollx = Scrollbar(emp_frame, orient=HORIZONTAL)

        self.EmployeeTable = ttk.Treeview(emp_frame,
                                          columns=("Emp-ID", "Name", "Email", "Gender", "CNIC",
                                                   "Contact", "DOB", "DOJ", "Address", "Salary"),
                                          yscrollcommand=scrolly.set,
                                          xscrollcommand=scrollx.set)

        # Style the treeview
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview",
                        background="white",
                        foreground="#333333",
                        fieldbackground="white",
                        rowheight=22,
                        font=("Segoe UI", 9))

        # Configure selected item colors - green selection
        style.map("Treeview",
                  background=[('selected', '#2ECC71')],
                  foreground=[('selected', 'white')])

        style.configure("Treeview.Heading",
                        background="#2ECC71",
                        foreground="white",
                        font=("Segoe UI", 9, "bold"),
                        relief=FLAT)
        style.map("Treeview.Heading",
                  background=[('active', '#009140')])

        scrollx.pack(side=BOTTOM, fill=X)
        scrolly.pack(side=RIGHT, fill=Y)
        scrollx.config(command=self.EmployeeTable.xview)
        scrolly.config(command=self.EmployeeTable.yview)

        self.EmployeeTable.heading("Emp-ID", text="Emp-ID")
        self.EmployeeTable.heading("Name", text="Name")
        self.EmployeeTable.heading("Email", text="Email")
        self.EmployeeTable.heading("Gender", text="Gender")
        self.EmployeeTable.heading("CNIC", text="CNIC")
        self.EmployeeTable.heading("Contact", text="Contact")
        self.EmployeeTable.heading("DOB", text="DOB")
        self.EmployeeTable.heading("DOJ", text="Joining Date")
        self.EmployeeTable.heading("Address", text="Address")
        self.EmployeeTable.heading("Salary", text="Salary")

        self.EmployeeTable["show"] = "headings"
        # Slightly reduced column widths for laptop
        self.EmployeeTable.column("Emp-ID", width=70, anchor=CENTER)
        self.EmployeeTable.column("Name", width=100, anchor=CENTER)
        self.EmployeeTable.column("Email", width=130, anchor=CENTER)
        self.EmployeeTable.column("Gender", width=70, anchor=CENTER)
        self.EmployeeTable.column("CNIC", width=120, anchor=CENTER)
        self.EmployeeTable.column("Contact", width=100, anchor=CENTER)
        self.EmployeeTable.column("DOB", width=85, anchor=CENTER)
        self.EmployeeTable.column("DOJ", width=85, anchor=CENTER)
        self.EmployeeTable.column("Address", width=150, anchor=CENTER)
        self.EmployeeTable.column("Salary", width=100, anchor=CENTER)

        self.EmployeeTable.pack(fill=BOTH, expand=True)
        self.EmployeeTable.bind("<ButtonRelease-1>", self.get_data)

        # Generate initial Emp-ID and show data
        self.generate_emp_id()
        self.show()

    # ---------- New methods for custom date entry ----------
    def create_date_entry(self, parent, textvariable, row, col, colspan=1):
        """Create a custom date entry with placeholder and calendar button"""
        frame = Frame(parent, bg="white", highlightthickness=1,
                      highlightbackground="#E0E6ED", highlightcolor="#2ECC71")
        frame.grid(row=row, column=col, columnspan=colspan, padx=(0, 10), pady=4, sticky='ew')

        # Entry for date
        entry = Entry(frame, font=("Segoe UI", 10), bg="white", fg="#999999",
                      relief=FLAT, bd=0)
        entry.pack(side=LEFT, fill=BOTH, expand=True, padx=(5, 0))

        # Placeholder text
        placeholder = "DD/MM/YYYY"
        entry.insert(0, placeholder)

        # Store placeholder and variable reference
        entry.placeholder = placeholder
        entry.textvariable = textvariable

        # Bind events
        entry.bind("<FocusIn>", self.on_date_focus_in)
        entry.bind("<FocusOut>", self.on_date_focus_out)
        entry.bind("<KeyRelease>", self.on_date_key_release)

        # Calendar button
        btn_cal = Button(frame, text="📅", font=("Segoe UI", 9),
                         bg="#2ECC71", fg="white", bd=0,
                         cursor="hand2", width=3,
                         command=lambda: self.open_calendar(entry, textvariable))
        btn_cal.pack(side=RIGHT, padx=(0, 2))

        return frame, entry

    def on_date_focus_in(self, event):
        """Clear placeholder on focus in"""
        entry = event.widget
        if entry.get() == entry.placeholder:
            entry.delete(0, END)
            entry.config(fg="#333333")

    def on_date_focus_out(self, event):
        """Restore placeholder if empty, otherwise validate"""
        entry = event.widget
        value = entry.get().strip()
        if not value or value == entry.placeholder:
            entry.delete(0, END)
            entry.insert(0, entry.placeholder)
            entry.config(fg="#999999")
            entry.textvariable.set("")
        else:
            # Validate date format
            if not self.validate_date_string(value):
                messagebox.showerror("Invalid Date",
                                     "Please enter a valid date in DD/MM/YYYY format.",
                                     parent=self.root)
                # Revert to placeholder
                entry.delete(0, END)
                entry.insert(0, entry.placeholder)
                entry.config(fg="#999999")
                entry.textvariable.set("")
            else:
                entry.textvariable.set(value)

    def on_date_key_release(self, event):
        """Format date as user types: only digits, auto-insert slashes"""
        entry = event.widget
        # Ignore if placeholder is showing
        if entry.get() == entry.placeholder:
            return

        # Get current text and remove any non-digit characters
        current = entry.get()
        digits = ''.join(filter(str.isdigit, current))

        # Limit to 8 digits (DDMMYYYY)
        if len(digits) > 8:
            digits = digits[:8]

        # Build formatted string with slashes
        formatted = ""
        if len(digits) > 0:
            formatted = digits[:2]
        if len(digits) > 2:
            formatted += "/" + digits[2:4]
        if len(digits) > 4:
            formatted += "/" + digits[4:8]

        # Update entry and preserve cursor position
        cursor_pos = entry.index(INSERT)
        entry.delete(0, END)
        entry.insert(0, formatted)

        # Adjust cursor to skip over slashes
        new_pos = cursor_pos
        if new_pos > 2:
            new_pos += 1          # after first slash
        if new_pos > 5:
            new_pos += 1          # after second slash
        entry.icursor(min(new_pos, len(formatted)))

        # Update the StringVar with the current formatted value
        entry.textvariable.set(formatted)

    def validate_date_string(self, date_str):
        """Check if date_str is a valid date in DD/MM/YYYY format"""
        if not date_str or date_str == "DD/MM/YYYY":
            return False
        try:
            datetime.strptime(date_str, "%d/%m/%Y")
            return True
        except ValueError:
            return False

    def open_calendar(self, entry, textvariable):
        """Open a calendar popup just below the entry field and set the selected date"""
        def set_date():
            selected = cal.selection_get()
            if selected:
                formatted = selected.strftime("%d/%m/%Y")
                entry.delete(0, END)
                entry.insert(0, formatted)
                entry.config(fg="#333333")
                textvariable.set(formatted)
            top.destroy()

        # Get the position of the entry widget
        x = entry.winfo_rootx()
        y = entry.winfo_rooty() + entry.winfo_height() + 5  # 5px below entry

        top = Toplevel(self.root)
        top.title("Select Date")
        top.grab_set()
        top.resizable(False, False)

        # Position the popup
        top.geometry(f"300x250+{x}+{y}")

        cal = Calendar(top, selectmode='day', date_pattern='dd/mm/yyyy')
        cal.pack(padx=10, pady=10)

        btn_ok = Button(top, text="OK", command=set_date,
                        bg="#2ECC71", fg="white", font=("Segoe UI", 9))
        btn_ok.pack(pady=5)

        # Ensure the popup stays on top
        top.transient(self.root)
        top.focus_force()
        top.grab_set()

    # ---------------------------------------------------------

    def update_datetime(self):
        """Update the real-time date and time label"""
        current_datetime = datetime.now().strftime("%A, %d %B %Y | %I:%M:%S %p")
        self.datetime_label.config(text=current_datetime)
        # Update every 1000 ms (1 second)
        self.root.after(1000, self.update_datetime)

    def on_close(self):
        """Handle window closing - return to dashboard"""
        if self.dashboard_window:
            self.dashboard_window.deiconify()  # Show dashboard window
        self.root.destroy()  # Close employee window

    def initialize_database(self):
        """Ensure the Employee table exists in the database with new structure"""
        try:
            with sqlite3.connect(database=r'Possystem.db') as con:
                cur = con.cursor()

                # Check if table exists
                cur.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name='Employee'")
                table_exists = cur.fetchone()

                if table_exists:
                    # Check table structure
                    cur.execute("PRAGMA table_info(Employee)")
                    columns = [col[1] for col in cur.fetchall()]

                    # Check for new columns
                    new_columns = ['ID_Card', 'Utility_Bill']
                    for new_col in new_columns:
                        if new_col not in columns:
                            cur.execute(
                                f"ALTER TABLE Employee ADD COLUMN {new_col} BLOB")
                else:
                    # Create table with new structure
                    cur.execute('''CREATE TABLE IF NOT EXISTS Employee (
                                EmpID TEXT PRIMARY KEY,
                                Name TEXT NOT NULL,
                                Email TEXT,
                                Gender TEXT,
                                CNIC TEXT UNIQUE,
                                Contact TEXT,
                                DOB TEXT,
                                DOJ TEXT,
                                Address TEXT,
                                Salary TEXT,
                                ID_Card BLOB,
                                Utility_Bill BLOB)''')

                con.commit()
        except Exception as e:
            print(f"Error initializing database: {e}")

    def upload_id_card(self):
        """Upload ID Card image"""
        file_types = [('Image files', '*.jpg *.jpeg *.png *.gif *.bmp')]
        file_path = filedialog.askopenfilename(title="Select ID Card Image",
                                               filetypes=file_types)
        if file_path:
            try:
                # Store the original path
                self.id_card_path = file_path
                # Read and store image data
                with open(file_path, 'rb') as f:
                    self.id_card_data = f.read()

                file_name = os.path.basename(file_path)
                self.lbl_id_status.config(
                    text=f"Uploaded: {file_name[:20]}...", fg="green")
                self.btn_view_id.config(state=NORMAL)
                messagebox.showinfo(
                    "Success", "ID Card uploaded successfully!", parent=self.root)
            except Exception as e:
                messagebox.showerror(
                    "Error", f"Error reading image: {str(e)}", parent=self.root)

    def upload_utility_bill(self):
        """Upload Utility Bill image"""
        file_types = [
            ('Image files', '*.jpg *.jpeg *.png *.gif *.bmp'), ('PDF files', '*.pdf')]
        file_path = filedialog.askopenfilename(title="Select Utility Bill",
                                               filetypes=file_types)
        if file_path:
            try:
                # Store the original path
                self.utility_bill_path = file_path
                # Read and store file data
                with open(file_path, 'rb') as f:
                    self.utility_bill_data = f.read()

                file_name = os.path.basename(file_path)
                self.lbl_utility_status.config(
                    text=f"Uploaded: {file_name[:20]}...", fg="green")
                self.btn_view_utility.config(state=NORMAL)
                messagebox.showinfo(
                    "Success", "Utility Bill uploaded successfully!", parent=self.root)
            except Exception as e:
                messagebox.showerror(
                    "Error", f"Error reading file: {str(e)}", parent=self.root)

    def view_id_card(self):
        """View uploaded ID Card"""
        if self.id_card_data:
            self.show_image(self.id_card_data, "ID Card", is_blob=True)
        else:
            messagebox.showwarning(
                "Warning", "No ID Card uploaded!", parent=self.root)

    def view_utility_bill(self):
        """View uploaded Utility Bill"""
        if self.utility_bill_data:
            # Check if it's a PDF
            if self.utility_bill_path.lower().endswith('.pdf'):
                # Save PDF to temp file and open
                temp_pdf = tempfile.NamedTemporaryFile(
                    delete=False, suffix='.pdf')
                temp_pdf.write(self.utility_bill_data)
                temp_pdf.close()

                import webbrowser
                webbrowser.open(temp_pdf.name)
                # Clean up temp file after some time
                self.root.after(5000, lambda: os.unlink(temp_pdf.name))
            else:
                self.show_image(self.utility_bill_data,
                                "Utility Bill", is_blob=True)
        else:
            messagebox.showwarning(
                "Warning", "No Utility Bill uploaded!", parent=self.root)

    def show_image(self, image_data, title, is_blob=False):
        """Show image in a new window - full screen"""
        try:
            img_window = Toplevel(self.root)
            img_window.title(title)

            # Get screen dimensions
            screen_width = img_window.winfo_screenwidth()
            screen_height = img_window.winfo_screenheight()

            # Set window to 90% of screen size
            window_width = int(screen_width * 0.9)
            window_height = int(screen_height * 0.9)

            # Center the window
            x_position = (screen_width - window_width) // 2
            y_position = (screen_height - window_height) // 2

            img_window.geometry(
                f"{window_width}x{window_height}+{x_position}+{y_position}")

            # Create scrollable canvas
            canvas = Canvas(img_window, bg="white")
            scrollbar_y = Scrollbar(
                img_window, orient=VERTICAL, command=canvas.yview)
            scrollbar_x = Scrollbar(
                img_window, orient=HORIZONTAL, command=canvas.xview)

            scrollable_frame = Frame(canvas, bg="white")

            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )

            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar_y.set,
                             xscrollcommand=scrollbar_x.set)

            # Load image
            if is_blob:
                img = PILImage.open(io.BytesIO(image_data))
            else:
                img = PILImage.open(image_data)

            # Calculate maximum size for display
            max_width = window_width - 50
            max_height = window_height - 100

            # Resize image while maintaining aspect ratio
            img_ratio = img.width / img.height
            if img.width > max_width or img.height > max_height:
                if img_ratio > 1:  # Wider than tall
                    new_width = max_width
                    new_height = int(max_width / img_ratio)
                else:  # Taller than wide
                    new_height = max_height
                    new_width = int(new_height * img_ratio)

                img = img.resize((new_width, new_height),
                                 PILImage.Resampling.LANCZOS)

            photo = ImageTk.PhotoImage(img)

            # Display image
            lbl_img = Label(scrollable_frame, image=photo, bg="white")
            lbl_img.image = photo  # Keep a reference
            lbl_img.pack(padx=20, pady=20)

            # Pack scrollbars and canvas
            scrollbar_y.pack(side=RIGHT, fill=Y)
            scrollbar_x.pack(side=BOTTOM, fill=X)
            canvas.pack(side=LEFT, fill=BOTH, expand=True)

            # Add zoom buttons
            zoom_frame = Frame(img_window, bg="white")
            zoom_frame.pack(side=BOTTOM, fill=X, pady=5)

            btn_zoom_in = Button(zoom_frame, text="Zoom In (+)",
                                 command=lambda: self.zoom_image(
                                     lbl_img, img, 1.2),
                                 font=("Segoe UI", 9), bg="#1A3C6E", fg="white")
            btn_zoom_in.pack(side=LEFT, padx=5)

            btn_zoom_out = Button(zoom_frame, text="Zoom Out (-)",
                                  command=lambda: self.zoom_image(
                                      lbl_img, img, 0.8),
                                  font=("Segoe UI", 9), bg="#1A3C6E", fg="white")
            btn_zoom_out.pack(side=LEFT, padx=5)

            btn_original = Button(zoom_frame, text="Original Size",
                                  command=lambda: self.reset_image(
                                      lbl_img, img),
                                  font=("Segoe UI", 9), bg="#2E8B57", fg="white")
            btn_original.pack(side=LEFT, padx=5)

            # Add close button
            btn_close = Button(zoom_frame, text="Close", command=img_window.destroy,
                               font=("Segoe UI", 10), bg="#C53030", fg="white")
            btn_close.pack(side=RIGHT, padx=5)

            # Enable mouse wheel scrolling
            def on_mouse_wheel(event):
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")

            img_window.bind_all("<MouseWheel>", on_mouse_wheel)

            # Clean up binding when window closes
            def on_close():
                img_window.unbind_all("<MouseWheel>")
                img_window.destroy()

            img_window.protocol("WM_DELETE_WINDOW", on_close)

        except Exception as e:
            messagebox.showerror(
                "Error", f"Cannot display image: {str(e)}", parent=self.root)

    def zoom_image(self, label, original_img, factor):
        """Zoom in/out on the image"""
        try:
            current_img = PILImage.open(io.BytesIO(self.id_card_data) if hasattr(
                self, 'id_card_data') else io.BytesIO(self.utility_bill_data))

            # Calculate new dimensions
            new_width = int(current_img.width * factor)
            new_height = int(current_img.height * factor)

            # Limit minimum and maximum size
            if new_width < 50 or new_height < 50:
                return
            if new_width > 5000 or new_height > 5000:
                return

            # Resize image
            resized_img = current_img.resize(
                (new_width, new_height), PILImage.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(resized_img)

            label.config(image=photo)
            label.image = photo  # Keep reference

        except Exception as e:
            print(f"Zoom error: {e}")

    def reset_image(self, label, original_img):
        """Reset image to original size"""
        try:
            photo = ImageTk.PhotoImage(original_img)
            label.config(image=photo)
            label.image = photo  # Keep reference
        except Exception as e:
            print(f"Reset error: {e}")

    def on_contact_focus(self, event):
        """Set cursor to the beginning of the entry when focused"""
        self.contact_entry.icursor(0)

    def format_salary(self, event=None):
        """Format salary with commas as user types"""
        current_value = self.salary_entry.get()

        if current_value:
            cleaned = ''
            decimal_count = 0
            for char in current_value:
                if char.isdigit():
                    cleaned += char
                elif char == '.' and decimal_count == 0:
                    cleaned += char
                    decimal_count += 1
                elif char == ',' or char == ' ':
                    continue
                else:
                    break

            if '.' in cleaned:
                parts = cleaned.split('.')
                if len(parts) > 1:
                    cleaned = parts[0] + '.' + parts[1][:2]

            if cleaned:
                self.salary_entry.delete(0, END)
                self.salary_entry.insert(0, cleaned)

    def format_salary_final(self, event=None):
        """Final formatting when focus leaves salary field"""
        current_value = self.salary_entry.get()
        if current_value:
            cleaned = current_value.replace(',', '').replace(' ', '')

            if cleaned:
                try:
                    if '.' in cleaned:
                        parts = cleaned.split('.')
                        integer_part = parts[0]
                        decimal_part = parts[1] if len(parts) > 1 else ''

                        if integer_part:
                            formatted_int = '{:,}'.format(int(integer_part))
                        else:
                            formatted_int = '0'

                        result = formatted_int
                        if decimal_part:
                            result += '.' + decimal_part
                    else:
                        result = '{:,}'.format(int(cleaned))

                    self.salary_entry.delete(0, END)
                    self.salary_entry.insert(0, result)
                except ValueError:
                    pass

    def generate_emp_id(self):
        """Generate a new unique Emp-ID starting with E followed by 4 digits"""
        try:
            with sqlite3.connect(database=r'Possystem.db') as con:
                cur = con.cursor()

                cur.execute('''CREATE TABLE IF NOT EXISTS Employee (
                            EmpID TEXT PRIMARY KEY,
                            Name TEXT NOT NULL,
                            Email TEXT,
                            Gender TEXT,
                            CNIC TEXT UNIQUE,
                            Contact TEXT,
                            DOB TEXT,
                            DOJ TEXT,
                            Address TEXT,
                            Salary TEXT,
                            ID_Card BLOB,
                            Utility_Bill BLOB)''')

                cur.execute(
                    "SELECT EmpID FROM employee WHERE EmpID LIKE 'E%' ORDER BY EmpID DESC LIMIT 1")
                result = cur.fetchone()

                if result:
                    max_id = result[0]
                    try:
                        num_part = int(max_id[1:])
                        new_num = num_part + 1
                    except ValueError:
                        new_num = 1
                else:
                    new_num = 1

                new_emp_id = f"E{new_num:04d}"
                self.var_EmpID.set(new_emp_id)

        except Exception as ex:
            messagebox.showerror(
                "Error", f"Error generating Emp-ID: {str(ex)}", parent=self.root)
            self.var_EmpID.set("E0001")

    def validate_cnic(self, event=None):
        """Format CNIC as XXXXX-XXXXXXX-X"""
        current_text = self.var_cnic.get()
        if current_text:
            digits = ''.join(filter(str.isdigit, current_text))
            if len(digits) > 13:
                digits = digits[:13]

            if len(digits) > 0:
                formatted = digits
                if len(digits) > 5:
                    formatted = digits[:5] + '-' + digits[5:]
                if len(digits) > 12:
                    formatted = digits[:5] + '-' + \
                        digits[5:12] + '-' + digits[12:]

                self.var_cnic.set(formatted)
                self.cnic_entry.icursor(len(formatted))

    def validate_contact(self, event=None):
        """Validate Contact to accept exactly 10 digits"""
        current_text = self.contact_entry.get()
        if current_text:
            digits = ''.join(filter(str.isdigit, current_text))
            if len(digits) > 10:
                digits = digits[:10]

            self.contact_entry.delete(0, END)
            self.contact_entry.insert(0, digits)
            self.contact_entry.icursor(len(digits))

    def validate_dates(self):
        """Validate that DOB is before DOJ and employee is at least 18"""
        try:
            dob_str = self.var_DOB.get()
            doj_str = self.var_DOJ.get()

            if dob_str and doj_str:
                dob_date = datetime.strptime(dob_str, "%d/%m/%Y")
                doj_date = datetime.strptime(doj_str, "%d/%m/%Y")

                if dob_date >= doj_date:
                    return False, "Date of Birth must be earlier than Joining Date"

                age_at_joining = doj_date.year - dob_date.year
                if age_at_joining < 18:
                    return False, "Employee must be at least 18 years old at joining date"

            return True, ""
        except ValueError:
            return True, ""

    def validate_kye_documents(self):
        """Validate that KYE documents are uploaded"""
        errors = []

        if not self.id_card_data:
            errors.append("ID Card must be uploaded")

        if not self.utility_bill_data:
            errors.append("Utility Bill must be uploaded")

        if errors:
            messagebox.showerror("KYE Validation Error",
                                 "\n".join(errors), parent=self.root)
            return False

        return True

    def validate_all_fields(self):
        """Validate that all required fields are filled"""
        errors = []

        # Check Emp-ID
        if not self.var_EmpID.get():
            errors.append("Employee ID is required")

        # Check Name
        if not self.var_name.get().strip():
            errors.append("Name is required")

        # Check Gender
        if self.var_gender.get() == "Select":
            errors.append("Please select Gender")

        # Check CNIC
        cnic = self.var_cnic.get().replace('-', '')
        if not cnic:
            errors.append("CNIC is required")
        elif len(cnic) != 13:
            errors.append("CNIC must be exactly 13 digits")

        # Check Contact
        contact = self.contact_entry.get()
        if not contact:
            errors.append("Contact number is required")
        elif len(contact) != 10:
            errors.append("Contact must be exactly 10 digits")
        elif not contact.isdigit():
            errors.append("Contact must contain only digits")

        # Check Email
        if not self.var_email.get():
            errors.append("Email is required")
        else:
            pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if not re.match(pattern, self.var_email.get()):
                errors.append("Please enter a valid email address")

        # Check DOB
        if not self.var_DOB.get():
            errors.append("Date of Birth is required")

        # Check DOJ
        if not self.var_DOJ.get():
            errors.append("Joining Date is required")

        # Validate dates
        if self.var_DOB.get() and self.var_DOJ.get():
            is_valid, msg = self.validate_dates()
            if not is_valid:
                errors.append(msg)

        # Check Address
        if not self.txt_address.get('1.0', END).strip():
            errors.append("Address is required")

        # Check Salary
        salary_value = self.salary_entry.get().replace(',', '').replace(' ', '')
        if not salary_value:
            errors.append("Salary is required")
        else:
            try:
                salary_val = float(salary_value)
                if salary_val <= 0:
                    errors.append("Salary must be greater than 0")
            except ValueError:
                errors.append("Salary must be a valid number")

        if errors:
            messagebox.showerror("Validation Error",
                                 "\n".join(errors), parent=self.root)
            return False

        # Validate KYE documents
        return self.validate_kye_documents()

    def add(self):
        if not self.validate_all_fields():
            return

        try:
            with sqlite3.connect(database=r'Possystem.db') as con:
                cur = con.cursor()

                # Check if EmpID already exists
                cur.execute("Select * from employee where EmpID=?",
                            (self.var_EmpID.get(),))
                row = cur.fetchone()
                if row is not None:
                    messagebox.showerror(
                        "Error", "This Employee ID already assigned, try different", parent=self.root)
                    return

                # Check if CNIC already exists
                cnic_clean = self.var_cnic.get().replace('-', '')
                cur.execute("Select * from employee where CNIC=?",
                            (cnic_clean,))
                row = cur.fetchone()
                if row is not None:
                    messagebox.showerror(
                        "Error", "This CNIC already exists in the system", parent=self.root)
                    return

                # Check if Email already exists
                cur.execute("Select * from employee where Email=?",
                            (self.var_email.get(),))
                row = cur.fetchone()
                if row is not None:
                    messagebox.showerror(
                        "Error", "This Email already exists in the system", parent=self.root)
                    return

                # Format contact with +92 prefix
                contact_full = "+92" + self.contact_entry.get()

                # Clean salary value
                salary_clean = self.salary_entry.get().replace(',', '').replace(' ', '')

                # Insert employee data with KYE documents as BLOB
                cur.execute("""Insert into Employee(EmpID,Name,Email,Gender,CNIC,Contact,DOB,DOJ,Address,Salary,ID_Card,Utility_Bill) 
                            values(?,?,?,?,?,?,?,?,?,?,?,?)""", (
                            self.var_EmpID.get(),
                            self.var_name.get().strip(),
                            self.var_email.get().strip(),
                            self.var_gender.get(),
                            cnic_clean,
                            contact_full,
                            self.var_DOB.get(),
                            self.var_DOJ.get(),
                            self.txt_address.get('1.0', END).strip(),
                            salary_clean,
                            self.id_card_data,
                            self.utility_bill_data
                            ))
                con.commit()
                messagebox.showinfo(
                    "Success", "Employee added successfully with KYE documents", parent=self.root)
                self.show()
                self.clear()

        except sqlite3.IntegrityError as e:
            if "UNIQUE" in str(e):
                if "CNIC" in str(e):
                    messagebox.showerror(
                        "Error", "CNIC already exists in the system", parent=self.root)
                elif "Email" in str(e):
                    messagebox.showerror(
                        "Error", "Email already exists in the system", parent=self.root)
            else:
                messagebox.showerror(
                    "Error", f"Database error: {str(e)}", parent=self.root)
        except Exception as ex:
            messagebox.showerror(
                "Error", f"Error due to: {str(ex)}", parent=self.root)

    def show(self):
        try:
            with sqlite3.connect(database=r'Possystem.db') as con:
                cur = con.cursor()
                cur.execute(
                    "SELECT EmpID,Name,Email,Gender,CNIC,Contact,DOB,DOJ,Address,Salary FROM employee ORDER BY EmpID")
                rows = cur.fetchall()

                self.EmployeeTable.delete(*self.EmployeeTable.get_children())

                for row in rows:
                    formatted_row = list(row)

                    # Format CNIC for display
                    if row[4] and len(str(row[4])) == 13:
                        cnic = str(row[4])
                        formatted_row[4] = f"{cnic[:5]}-{cnic[5:12]}-{cnic[12:]}"

                    # Format salary with Rs prefix and commas
                    if row[9]:
                        try:
                            salary_value = float(row[9])
                            formatted_row[9] = "Rs " + \
                                '{:,.2f}'.format(salary_value)
                        except (ValueError, TypeError):
                            formatted_row[9] = row[9]

                    self.EmployeeTable.insert('', END, values=formatted_row)

        except Exception as ex:
            messagebox.showerror(
                "Error", f"Error displaying data: {str(ex)}", parent=self.root)

    def get_data(self, ev):
        f = self.EmployeeTable.focus()
        if not f:
            return

        content = self.EmployeeTable.item(f)
        row = content['values']
        if not row:
            return

        try:
            self.var_EmpID.set(row[0])
            self.var_name.set(row[1])
            self.var_email.set(row[2])
            self.var_gender.set(row[3])

            # Set CNIC
            cnic = str(row[4])
            self.var_cnic.set(cnic)

            # Set contact
            contact = str(row[5])
            if contact.startswith('+92'):
                contact = contact[3:]
            elif contact.startswith('92') and len(contact) == 12:
                contact = contact[2:]

            if len(contact) == 10 and contact.isdigit():
                self.contact_entry.delete(0, END)
                self.contact_entry.insert(0, contact)
            else:
                digits = ''.join(filter(str.isdigit, contact))
                if len(digits) >= 10:
                    digits = digits[-10:]
                self.contact_entry.delete(0, END)
                self.contact_entry.insert(0, digits)

            # Set dates in custom entries
            dob = row[6] if row[6] else ""
            doj = row[7] if row[7] else ""

            # Update DOB entry
            self.dob_entry.delete(0, END)
            if dob:
                self.dob_entry.insert(0, dob)
                self.dob_entry.config(fg="#333333")
                self.var_DOB.set(dob)
            else:
                self.dob_entry.insert(0, self.dob_entry.placeholder)
                self.dob_entry.config(fg="#999999")
                self.var_DOB.set("")

            # Update DOJ entry
            self.doj_entry.delete(0, END)
            if doj:
                self.doj_entry.insert(0, doj)
                self.doj_entry.config(fg="#333333")
                self.var_DOJ.set(doj)
            else:
                self.doj_entry.insert(0, self.doj_entry.placeholder)
                self.doj_entry.config(fg="#999999")
                self.var_DOJ.set("")

            # Set Address
            self.txt_address.delete('1.0', END)
            self.txt_address.insert(END, row[8])

            # Set salary
            salary = str(row[9])
            if salary.startswith('Rs'):
                salary = salary[2:].strip()
            self.salary_entry.delete(0, END)
            self.salary_entry.insert(0, salary)

            # Load KYE documents from database
            self.load_kye_documents()

            # Make Emp-ID editable for update
            self.txt_empid.config(state='normal')

        except IndexError as e:
            messagebox.showerror(
                "Error", f"Error loading data: {str(e)}", parent=self.root)
        except Exception as e:
            messagebox.showerror(
                "Error", f"Error loading employee data: {str(e)}", parent=self.root)

    def load_kye_documents(self):
        """Load KYE document data from database"""
        try:
            with sqlite3.connect(database=r'Possystem.db') as con:
                cur = con.cursor()
                cur.execute(
                    "SELECT ID_Card, Utility_Bill FROM employee WHERE EmpID=?", (self.var_EmpID.get(),))
                row = cur.fetchone()

                if row:
                    id_card_data, utility_bill_data = row

                    # Update ID Card status
                    if id_card_data:
                        self.id_card_data = id_card_data
                        self.lbl_id_status.config(
                            text="Loaded: ID Card from database", fg="green")
                        self.btn_view_id.config(state=NORMAL)
                    else:
                        self.id_card_data = None
                        self.lbl_id_status.config(
                            text="No file selected", fg="red")
                        self.btn_view_id.config(state=DISABLED)

                    # Update Utility Bill status
                    if utility_bill_data:
                        self.utility_bill_data = utility_bill_data
                        self.lbl_utility_status.config(
                            text="Loaded: Utility Bill from database", fg="green")
                        self.btn_view_utility.config(state=NORMAL)
                    else:
                        self.utility_bill_data = None
                        self.lbl_utility_status.config(
                            text="No file selected", fg="red")
                        self.btn_view_utility.config(state=DISABLED)
        except Exception as e:
            print(f"Error loading KYE documents: {e}")

    def update(self):
        if not self.validate_all_fields():
            return

        try:
            with sqlite3.connect(database=r'Possystem.db') as con:
                cur = con.cursor()

                # Check if EmpID exists
                cur.execute("Select * from employee where EmpID=?",
                            (self.var_EmpID.get(),))
                row = cur.fetchone()
                if row is None:
                    messagebox.showerror(
                        "Error", "Invalid Employee ID", parent=self.root)
                    return

                # Check if CNIC already exists for another employee
                cnic_clean = self.var_cnic.get().replace('-', '')
                cur.execute("Select * from employee where CNIC=? AND EmpID!=?",
                            (cnic_clean, self.var_EmpID.get()))
                row = cur.fetchone()
                if row is not None:
                    messagebox.showerror(
                        "Error", "This CNIC already exists for another employee", parent=self.root)
                    return

                # Check if Email already exists for another employee
                cur.execute("Select * from employee where Email=? AND EmpID!=?",
                            (self.var_email.get(), self.var_EmpID.get()))
                row = cur.fetchone()
                if row is not None:
                    messagebox.showerror(
                        "Error", "This Email already exists for another employee", parent=self.root)
                    return

                # Format contact with +92 prefix
                contact_full = "+92" + self.contact_entry.get()

                # Clean salary value
                salary_clean = self.salary_entry.get().replace(',', '').replace(' ', '')

                # Update employee data with KYE documents as BLOB
                cur.execute("""Update employee set Name=?,Email=?,Gender=?,CNIC=?,Contact=?,DOB=?,DOJ=?,Address=?,Salary=?,ID_Card=?,Utility_Bill=? 
                            where EmpID=?""", (
                            self.var_name.get().strip(),
                            self.var_email.get().strip(),
                            self.var_gender.get(),
                            cnic_clean,
                            contact_full,
                            self.var_DOB.get(),
                            self.var_DOJ.get(),
                            self.txt_address.get('1.0', END).strip(),
                            salary_clean,
                            self.id_card_data,
                            self.utility_bill_data,
                            self.var_EmpID.get(),
                            ))
                con.commit()

                if cur.rowcount > 0:
                    messagebox.showinfo(
                        "Success", "Employee updated successfully with KYE documents", parent=self.root)
                    self.show()
                    self.txt_empid.config(state='readonly')
                else:
                    messagebox.showerror(
                        "Error", "No changes made to employee record", parent=self.root)

        except sqlite3.IntegrityError as e:
            if "UNIQUE" in str(e):
                messagebox.showerror(
                    "Error", "Duplicate entry. CNIC or Email already exists for another employee.", parent=self.root)
            else:
                messagebox.showerror(
                    "Error", f"Database error: {str(e)}", parent=self.root)
        except Exception as ex:
            messagebox.showerror(
                "Error", f"Error updating employee: {str(ex)}", parent=self.root)

    def delete(self):
        if not self.var_EmpID.get():
            messagebox.showerror(
                "Error", "Please select an employee to delete", parent=self.root)
            return

        try:
            with sqlite3.connect(database=r'Possystem.db') as con:
                cur = con.cursor()

                op = messagebox.askyesno(
                    "Confirm", "Do you really want to delete this employee?", parent=self.root)
                if op:
                    cur.execute("DELETE FROM employee WHERE EmpID=?",
                                (self.var_EmpID.get(),))
                    con.commit()

                    if cur.rowcount > 0:
                        messagebox.showinfo(
                            "Delete", "Employee deleted successfully", parent=self.root)
                        self.clear()
                    else:
                        messagebox.showerror(
                            "Error", "Employee not found", parent=self.root)

        except Exception as ex:
            messagebox.showerror(
                "Error", f"Error deleting employee: {str(ex)}", parent=self.root)

    def clear(self):
        self.var_name.set("")
        self.var_email.set("")
        self.var_gender.set("Select")
        self.var_cnic.set("")
        self.contact_entry.delete(0, END)
        # Reset date entries to placeholder
        for entry in [self.dob_entry, self.doj_entry]:
            entry.delete(0, END)
            entry.insert(0, entry.placeholder)
            entry.config(fg="#999999")
        self.var_DOB.set("")
        self.var_DOJ.set("")
        self.txt_address.delete('1.0', END)
        self.salary_entry.delete(0, END)
        self.var_searchtxt.set("")
        self.var_searchby.set("Select")

        # Clear KYE documents
        self.id_card_path = ""
        self.utility_bill_path = ""
        self.id_card_data = None
        self.utility_bill_data = None
        self.lbl_id_status.config(text="No file selected", fg="red")
        self.lbl_utility_status.config(text="No file selected", fg="red")
        self.btn_view_id.config(state=DISABLED)
        self.btn_view_utility.config(state=DISABLED)

        # Generate new Emp-ID
        self.generate_emp_id()
        self.show()
        self.txt_empid.config(state='readonly')

    def search(self):
        valid_columns = ["EmpID", "Name", "Email", "Gender",
                         "CNIC", "Contact", "DOB", "DOJ", "Address", "Salary"]

        if self.var_searchby.get() == "Select":
            messagebox.showerror(
                "Error", "Select Search by option", parent=self.root)
            return

        if not self.var_searchtxt.get().strip():
            messagebox.showerror(
                "Error", "Search input is required", parent=self.root)
            return

        search_column = self.var_searchby.get()
        search_text = self.var_searchtxt.get().strip()

        # Map display name to database column name
        if search_column == "Emp-ID":
            search_column = "EmpID"

        if search_column not in valid_columns:
            messagebox.showerror(
                "Error", "Invalid search column", parent=self.root)
            return

        try:
            with sqlite3.connect(database=r'Possystem.db') as con:
                cur = con.cursor()

                # Prepare search text based on column
                if search_column == "EmpID" and not search_text.startswith("E"):
                    search_text = "E" + search_text
                elif search_column == "CNIC":
                    search_text = search_text.replace("-", "")
                elif search_column == "Contact":
                    if not search_text.startswith('+92'):
                        search_text = "+92" + search_text

                query = f"SELECT EmpID,Name,Email,Gender,CNIC,Contact,DOB,DOJ,Address,Salary FROM employee WHERE {search_column} LIKE ?"
                cur.execute(query, ("%" + search_text + "%",))
                rows = cur.fetchall()

                self.EmployeeTable.delete(*self.EmployeeTable.get_children())

                if rows:
                    for row in rows:
                        formatted_row = list(row)
                        if row[4] and len(str(row[4])) == 13:
                            cnic = str(row[4])
                            formatted_row[4] = f"{cnic[:5]}-{cnic[5:12]}-{cnic[12:]}"
                        if row[9]:
                            try:
                                salary_value = float(row[9])
                                formatted_row[9] = "Rs " + \
                                    '{:,.2f}'.format(salary_value)
                            except:
                                pass
                        self.EmployeeTable.insert(
                            "", END, values=formatted_row)
                else:
                    messagebox.showinfo(
                        "Result", "No records found", parent=self.root)

        except Exception as ex:
            messagebox.showerror(
                "Error", f"Error searching: {str(ex)}", parent=self.root)

    def add_pdf_footer(self, pdf):
        """Add footer to PDF page with company information"""
        # Save current position
        current_y = pdf.get_y()

        # Add footer separator
        pdf.set_draw_color(200, 200, 200)
        pdf.line(10, 280, 280, 280)

        # Move to footer position
        pdf.set_y(285)

        # Set footer font and color
        pdf.set_font('Arial', 'I', 8)
        pdf.set_text_color(100, 100, 100)

        # Add footer text
        footer_text = "Software produced by Dukes Tech Services | Website: dukestechservices.com | Phone: +923097671363"
        pdf.cell(0, 4, footer_text, ln=True, align='C')

        # Add page number
        pdf.cell(0, 4, f"Page {pdf.page_no()}", ln=True, align='C')

        # Reset position if needed
        pdf.set_y(current_y)

    def generate_pdf_report(self):
        """Generate professional PDF report of all employees"""
        try:
            # Fetch data from database
            with sqlite3.connect(database=r'Possystem.db') as con:
                cur = con.cursor()
                cur.execute(
                    "SELECT EmpID,Name,Email,Gender,CNIC,Contact,DOB,DOJ,Address,Salary FROM employee ORDER BY EmpID")
                rows = cur.fetchall()

                if not rows:
                    messagebox.showinfo(
                        "Info", "No employee records found", parent=self.root)
                    return

                # Ask for save location
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".pdf",
                    filetypes=[("PDF files", "*.pdf")],
                    title="Save PDF Report As",
                    initialfile="Employee_Report.pdf"
                )

                if not file_path:
                    return

                # Create PDF with professional styling
                pdf = FPDF(orientation='L')  # Landscape for better table fit
                pdf.add_page()

                # Set fonts
                pdf.add_font(
                    'Arial', '', r'C:\Windows\Fonts\arial.ttf', uni=True)
                pdf.add_font(
                    'Arial', 'B', r'C:\Windows\Fonts\arialbd.ttf', uni=True)

                # Title
                pdf.set_font('Arial', 'B', 20)
                pdf.set_text_color(11, 93, 30)  # Blue color
                pdf.cell(0, 15, "EMPLOYEE HISTORICAL REPORT",
                         ln=True, align='C')

                # Subtitle
                pdf.set_font('Arial', '', 12)
                pdf.set_text_color(100, 100, 100)
                pdf.cell(
                    0, 8, f"Generated on: {datetime.now().strftime('%d %B, %Y at %I:%M %p')}", ln=True, align='C')

                # Company info
                pdf.set_font('Arial', 'I', 10)
                pdf.cell(
                    0, 8, "Inventory Management System | Developed by Dukes Tech Services", ln=True, align='C')

                pdf.ln(15)

                # ===== STATISTICS SUMMARY AT THE START =====
                pdf.set_font('Arial', 'B', 14)
                pdf.set_text_color(11, 93, 30)
                pdf.cell(0, 10, "Statistics Summary:", ln=True)

                # Calculate statistics
                male_count = sum(1 for row in rows if str(
                    row[3]).lower() == 'male')
                female_count = sum(1 for row in rows if str(
                    row[3]).lower() == 'female')
                other_count = sum(1 for row in rows if str(
                    row[3]).lower() not in ['male', 'female'])

                try:
                    total_salary = sum(float(row[9]) for row in rows if row[9])
                    avg_salary = total_salary / len(rows) if len(rows) > 0 else 0
                except:
                    total_salary = 0
                    avg_salary = 0

                pdf.set_font('Arial', '', 11)
                pdf.set_text_color(0, 0, 0)
                pdf.cell(0, 8, f"• Total Employees: {len(rows)}", ln=True)
                pdf.cell(0, 8, f"• Male Employees: {male_count}", ln=True)
                pdf.cell(0, 8, f"• Female Employees: {female_count}", ln=True)
                pdf.cell(0, 8, f"• Other Gender: {other_count}", ln=True)
                pdf.cell(
                    0, 8, f"• Total Monthly Salary: Rs {total_salary:,.2f}", ln=True)
                pdf.cell(
                    0, 8, f"• Average Salary: Rs {avg_salary:,.2f}", ln=True)

                pdf.ln(15)

                # Table header
                pdf.set_fill_color(11, 93, 30)  # Header background
                pdf.set_text_color(255, 255, 255)  # White text
                pdf.set_font('Arial', 'B', 10)

                # Column widths for landscape
                col_widths = [15, 30, 35, 15, 30, 25, 20, 20, 45, 25]

                headers = ["Emp-ID", "Name", "Email", "Gender",
                           "CNIC", "Contact", "DOB", "DOJ", "Address", "Salary"]

                for i, header in enumerate(headers):
                    pdf.cell(col_widths[i], 10, header,
                             border=1, fill=True, align='C')
                pdf.ln()

                # Table rows
                # Light gray for alternate rows
                pdf.set_fill_color(245, 247, 250)
                pdf.set_text_color(0, 0, 0)  # Black text
                pdf.set_font('Arial', '', 9)

                fill = False
                row_count = 0
                rows_per_page = 20  # Number of rows per page

                for row_idx, row in enumerate(rows):
                    # Check if we need a new page
                    if row_count > 0 and row_count % rows_per_page == 0:
                        # Add footer to the current page before adding new page
                        self.add_pdf_footer(pdf)

                        # Add new page
                        pdf.add_page()

                        # Re-add title on new page
                        pdf.set_font('Arial', 'B', 16)
                        pdf.set_text_color(11, 93, 30)
                        pdf.cell(0, 10, "EMPLOYEE HISTORICAL REPORT (Continued)",
                                 ln=True, align='C')
                        pdf.ln(5)

                        # Re-add table header on new page
                        pdf.set_fill_color(11, 93, 30)
                        pdf.set_text_color(255, 255, 255)
                        pdf.set_font('Arial', 'B', 10)
                        for i, header in enumerate(headers):
                            pdf.cell(col_widths[i], 10, header,
                                     border=1, fill=True, align='C')
                        pdf.ln()

                        # Reset row color
                        fill = False
                        pdf.set_font('Arial', '', 9)

                    # Alternate row colors
                    if fill:
                        pdf.set_fill_color(232, 245, 233)
                    else:
                        pdf.set_fill_color(255, 255, 255)

                    formatted_row = list(row)

                    # Format CNIC
                    if row[4] and len(str(row[4])) == 13:
                        cnic = str(row[4])
                        formatted_row[4] = f"{cnic[:5]}-{cnic[5:12]}-{cnic[12:]}"

                    # Format salary
                    if row[9]:
                        try:
                            salary_value = float(row[9])
                            formatted_row[9] = 'Rs {:,.2f}'.format(
                                salary_value)
                        except:
                            formatted_row[9] = row[9]

                    # Format other fields for better readability
                    if len(formatted_row[8]) > 30:  # Address
                        formatted_row[8] = formatted_row[8][:27] + "..."

                    if len(formatted_row[2]) > 25:  # Email
                        formatted_row[2] = formatted_row[2][:22] + "..."

                    if len(formatted_row[1]) > 20:  # Name
                        formatted_row[1] = formatted_row[1][:17] + "..."

                    # Add row cells
                    for i, item in enumerate(formatted_row):
                        if i == 9:  # Salary column - right align
                            pdf.cell(col_widths[i], 8, str(
                                item), border=1, fill=True, align='R')
                        else:
                            pdf.cell(col_widths[i], 8, str(
                                item), border=1, fill=True, align='L')
                    pdf.ln()
                    fill = not fill
                    row_count += 1

                # Add footer to the last page
                self.add_pdf_footer(pdf)

                # Save PDF
                pdf.output(file_path)
                messagebox.showinfo(
                    "Success", f"Professional PDF report generated successfully!\nSaved to: {file_path}", parent=self.root)

        except Exception as e:
            messagebox.showerror(
                "Error", f"Error generating PDF report: {str(e)}", parent=self.root)

    def generate_excel_report(self):
        """Generate Excel report of all employees"""
        try:
            # Fetch data from database
            with sqlite3.connect(database=r'Possystem.db') as con:
                cur = con.cursor()
                cur.execute(
                    "SELECT EmpID as 'Emp-ID', Name, Email, Gender, CNIC, Contact, DOB, DOJ, Address, Salary FROM employee ORDER BY EmpID")
                rows = cur.fetchall()

                if not rows:
                    messagebox.showinfo(
                        "Info", "No employee records found", parent=self.root)
                    return

                # Get column names
                cur.execute("PRAGMA table_info(employee)")
                columns_info = cur.fetchall()
                column_names = [info[1] for info in columns_info if info[1]
                                != 'ID_Card' and info[1] != 'Utility_Bill']

                # Create DataFrame
                df = pd.DataFrame(rows, columns=column_names)

                # Format CNIC
                df['CNIC'] = df['CNIC'].apply(
                    lambda x: f"{x[:5]}-{x[5:12]}-{x[12:]}" if pd.notnull(x) and len(str(x)) == 13 else x)

                # Format Salary
                df['Salary'] = df['Salary'].apply(lambda x: f"Rs {float(x):,.2f}" if pd.notnull(
                    x) and str(x).replace('.', '').isdigit() else x)

                # Ask for save location
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx"),
                               ("Excel 97-2003", "*.xls")],
                    title="Save Excel Report As",
                    initialfile="Employee_Report.xlsx"
                )

                if file_path:
                    # Create Excel writer
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        df.to_excel(
                            writer, sheet_name='Employees', index=False)

                        # Get workbook and worksheet for styling
                        workbook = writer.book
                        worksheet = writer.sheets['Employees']

                        # Apply professional styling
                        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                        from openpyxl.utils import get_column_letter

                        # Header styling
                        header_fill = PatternFill(
                            start_color="0B5D1E", end_color="0B5D1E", fill_type="solid")
                        header_font = Font(color="FFFFFF", bold=True, size=11)
                        header_alignment = Alignment(
                            horizontal="center", vertical="center")

                        # Data styling
                        data_font = Font(size=10)
                        data_alignment = Alignment(
                            horizontal="left", vertical="center")
                        salary_alignment = Alignment(
                            horizontal="right", vertical="center")

                        # Border
                        thin_border = Border(left=Side(style='thin'),
                                             right=Side(style='thin'),
                                             top=Side(style='thin'),
                                             bottom=Side(style='thin'))

                        # Apply styles to header
                        for col in range(1, len(column_names) + 1):
                            cell = worksheet.cell(row=1, column=col)
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = header_alignment
                            cell.border = thin_border

                        # Apply styles to data
                        for row in range(2, len(df) + 2):
                            # Alternate row colors
                            if row % 2 == 0:
                                fill_color = "E8F5E9"  # Light gray
                            else:
                                fill_color = "FFFFFF"  # White

                            row_fill = PatternFill(
                                start_color=fill_color, end_color=fill_color, fill_type="solid")

                            for col in range(1, len(column_names) + 1):
                                cell = worksheet.cell(row=row, column=col)
                                cell.font = data_font
                                cell.border = thin_border
                                cell.fill = row_fill

                                # Special alignment for salary column
                                if col == len(column_names):  # Last column (Salary)
                                    cell.alignment = salary_alignment
                                    cell.font = Font(size=10, bold=True, color="0B5D1E")
                                else:
                                    cell.alignment = data_alignment

                        # Auto-adjust column widths
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 3, 40)
                            worksheet.column_dimensions[column_letter].width = adjusted_width

                        # Add title
                        worksheet.insert_rows(1)
                        worksheet.merge_cells('A1:J1')
                        title_cell = worksheet['A1']
                        title_cell.value = "EMPLOYEE HISTORICAL REPORT"
                        title_cell.font = Font(
                            size=16, bold=True, color="0B5D1E")
                        title_cell.alignment = Alignment(
                            horizontal="center", vertical="center")

                        # Add subtitle
                        worksheet.insert_rows(2)
                        worksheet.merge_cells('A2:J2')
                        subtitle_cell = worksheet['A2']
                        subtitle_cell.value = f"Generated on: {datetime.now().strftime('%d %B, %Y at %I:%M %p')}"
                        subtitle_cell.font = Font(
                            size=10, italic=True, color="666666")
                        subtitle_cell.alignment = Alignment(
                            horizontal="center", vertical="center")

                        # Add footer with company info
                        worksheet.insert_rows(len(df) + 5)
                        worksheet.merge_cells(f'A{len(df) + 5}:J{len(df) + 5}')
                        footer_cell = worksheet[f'A{len(df) + 5}']
                        footer_cell.value = "Software produced by Dukes Tech Services | Website: dukestechservices.com | Phone: +923097671363"
                        footer_cell.font = Font(
                            size=9, italic=True, color="138808")
                        footer_cell.alignment = Alignment(
                            horizontal="center", vertical="center")

                        # Add confidentiality notice
                        worksheet.insert_rows(len(df) + 6)
                        worksheet.merge_cells(f'A{len(df) + 6}:J{len(df) + 6}')
                        notice_cell = worksheet[f'A{len(df) + 6}']
                        notice_cell.value = "Confidential Report - For Internal Use Only"
                        notice_cell.font = Font(
                            size=8, italic=True, color="999999")
                        notice_cell.alignment = Alignment(
                            horizontal="center", vertical="center")

                    messagebox.showinfo(
                        "Success", f"Professional Excel report generated successfully!\nSaved to: {file_path}", parent=self.root)

        except Exception as e:
            messagebox.showerror(
                "Error", f"Error generating Excel report: {str(e)}", parent=self.root)


if __name__ == "__main__":
    root = Tk()
    obj = EmployeeClass(root)
    root.mainloop()