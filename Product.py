from tkinter import *
from tkinter import ttk, messagebox, filedialog
from PIL import Image, ImageTk
import sqlite3
import locale
import time
from datetime import datetime
import os
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.lib.units import inch
import webbrowser
import pandas as pd
import openpyxl
import traceback

class IntegratedInventorySystem:
    def __init__(self, root):
        self.root = root
        self.root.geometry("1400x700+50+50")
        self.root.title("Inventory Management System | Developed by Dukes Tech Services")
        self.root.config(bg="#ffffff")
        self.root.focus_force()
        
        # Make window responsive
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        
        # ===== THEME COLOR =====
        self.primary_color = "#9B59B6"
        self.primary_light = "#B07AC0"
        self.primary_dark = "#6C3483"
        self.accent_color = "#F1C40F"
        self.success_color = "#2ECC71"
        self.danger_color = "#E74C3C"
        self.secondary_color = "#95A5A6"
        
        # BUTTON COLORS
        self.btn_save_color = "#00C853"
        self.btn_save_hover = "#00B248"
        self.btn_update_color = "#2962FF"
        self.btn_update_hover = "#1E4ED8"
        self.btn_delete_color = "#FF1744"
        self.btn_delete_hover = "#D50000"
        self.btn_clear_color = "#FF9100"
        self.btn_clear_hover = "#FF6D00"
        self.btn_pdf_color = "#E53935"
        self.btn_pdf_hover = "#C62828"
        # =========================
        
        # ========== Variables for Category ==========
        self.var_CatID = StringVar()
        self.var_CatName = StringVar()
        
        # ========== Variables for Product ==========
        self.var_pid = StringVar()
        self.var_ProdCat = StringVar()
        self.var_prod_name = StringVar()
        self.var_barcode = StringVar()          # Barcode field (optional)
        self.var_cost_price = StringVar()
        self.var_price = StringVar()
        self.var_qty = StringVar()
        self.var_status = StringVar()
        self.var_margin = StringVar()
        self.var_searchby = StringVar()
        self.var_searchtxt = StringVar()
        
        # Price formatting variable
        self.price_without_format = ""
        self.cost_price_without_format = ""
        
        # Reorder threshold
        self.reorder_threshold = 5
        
        # Lists for dropdowns
        self.cat_list = []
        
        # ========== Main Container ==========
        main_container = Frame(self.root, bg="white")
        main_container.grid(row=0, column=0, sticky="nsew")
        main_container.grid_rowconfigure(0, weight=0)
        main_container.grid_rowconfigure(1, weight=1)
        main_container.grid_columnconfigure(0, weight=1)
        
        # ========== Header Frame with Date & Time ==========
        header_frame = Frame(main_container, bg=self.primary_color, bd=0, relief=FLAT)
        header_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=5)
        header_frame.grid_columnconfigure(0, weight=1)
        
        lbl_title = Label(header_frame, text="Inventory Management System", 
                         font=("Segoe UI", 22, "bold"), bg=self.primary_color, 
                         fg="white", bd=0)
        lbl_title.grid(row=0, column=0, sticky="w", padx=15, pady=8)
        
        self.lbl_datetime = Label(header_frame, text="", 
                                 font=("Segoe UI", 10), bg=self.primary_color, 
                                 fg="#f0e6ff")
        self.lbl_datetime.grid(row=0, column=1, sticky="e", padx=20, pady=8)
        
        self.update_datetime()
        
        # ========== Main Content Frame (Left + Right) ==========
        content_frame = Frame(main_container, bg="white")
        content_frame.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        content_frame.grid_rowconfigure(0, weight=1)
        content_frame.grid_columnconfigure(0, weight=1, minsize=400)
        content_frame.grid_columnconfigure(1, weight=2, minsize=700)
        
        # ========== LEFT FRAME: Categories ==========
        left_frame = Frame(content_frame, bg="white", bd=1, relief=SOLID, highlightbackground="#dddddd", highlightthickness=1)
        left_frame.grid(row=0, column=0, sticky="nsew", padx=(0,2), pady=0)
        left_frame.grid_rowconfigure(4, weight=1)
        left_frame.grid_columnconfigure(0, weight=1)
        
        # Category title
        cat_title = Label(left_frame, text="Manage Categories", 
                         font=("Segoe UI", 14, "bold"), 
                         bg=self.primary_color, fg="white", padx=10, pady=5)
        cat_title.grid(row=0, column=0, sticky="ew")
        
        # Category entry frame
        cat_entry_frame = Frame(left_frame, bg="white", padx=10, pady=10)
        cat_entry_frame.grid(row=1, column=0, sticky="ew")
        cat_entry_frame.grid_columnconfigure(1, weight=1)
        
        lbl_cat_name = Label(cat_entry_frame, text="Category Name", 
                            font=("Segoe UI", 10), bg="white", fg="#2C3E50")
        lbl_cat_name.grid(row=0, column=0, sticky=W, pady=5)
        
        txt_cat_name = Entry(cat_entry_frame, textvariable=self.var_CatName, 
                            font=("Segoe UI", 10), bg="#F8F9F9", 
                            fg="#2C3E50", bd=1, relief=SOLID)
        txt_cat_name.grid(row=0, column=1, pady=5, padx=5, sticky="ew")
        
        # Category buttons frame (uniform width)
        cat_btn_frame = Frame(left_frame, bg="white", padx=8, pady=8)
        cat_btn_frame.grid(row=2, column=0, sticky="ew")
        for i in range(3):
            cat_btn_frame.grid_columnconfigure(i, weight=1, uniform="cat_btn")
        
        btn_add_cat = Button(cat_btn_frame, text="ADD", command=self.add_category,
                           font=("Segoe UI", 9, "bold"), bg=self.primary_color, 
                           fg="white", cursor="hand2", bd=0,
                           activebackground=self.primary_light)
        btn_add_cat.grid(row=0, column=0, padx=3, pady=3, sticky="ew")
        
        btn_delete_cat = Button(cat_btn_frame, text="DELETE", command=self.delete_category,
                              font=("Segoe UI", 9, "bold"), bg=self.btn_delete_color, 
                              fg="white", cursor="hand2", bd=0,
                              activebackground=self.btn_delete_hover)
        btn_delete_cat.grid(row=0, column=1, padx=3, pady=3, sticky="ew")
        
        btn_clear_cat = Button(cat_btn_frame, text="CLEAR", command=self.clear_category,
                             font=("Segoe UI", 9, "bold"), bg=self.btn_clear_color, 
                             fg="white", cursor="hand2", bd=0,
                             activebackground=self.btn_clear_hover)
        btn_clear_cat.grid(row=0, column=2, padx=3, pady=3, sticky="ew")
        
        # Category list frame
        cat_list_frame = Frame(left_frame, bg="white", bd=1, relief=SOLID)
        cat_list_frame.grid(row=4, column=0, sticky="nsew", padx=8, pady=8)
        cat_list_frame.grid_rowconfigure(0, weight=1)
        cat_list_frame.grid_columnconfigure(0, weight=1)
        
        scrolly_cat = Scrollbar(cat_list_frame, orient=VERTICAL)
        scrollx_cat = Scrollbar(cat_list_frame, orient=HORIZONTAL)
        
        self.Category_Table = ttk.Treeview(cat_list_frame, columns=("CID", "Name"),
                                          yscrollcommand=scrolly_cat.set,
                                          xscrollcommand=scrollx_cat.set)
        
        scrollx_cat.pack(side=BOTTOM, fill=X)
        scrolly_cat.pack(side=RIGHT, fill=Y)
        scrollx_cat.config(command=self.Category_Table.xview)
        scrolly_cat.config(command=self.Category_Table.yview)
        
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", 
                       background="white",
                       foreground="#2C3E50",
                       rowheight=25,
                       fieldbackground="white",
                       font=("Segoe UI", 9))
        style.configure("Treeview.Heading", 
                       background=self.primary_color,
                       foreground="white",
                       font=("Segoe UI", 9, "bold"))
        style.map("Treeview", background=[('selected', self.primary_light)])
        
        self.Category_Table.heading("CID", text="ID")
        self.Category_Table.heading("Name", text="Category Name")
        self.Category_Table["show"] = "headings"
        self.Category_Table.column("CID", width=40, anchor=CENTER)
        self.Category_Table.column("Name", width=150)
        self.Category_Table.pack(fill=BOTH, expand=True)
        self.Category_Table.bind("<ButtonRelease-1>", self.get_category_data)
        
        # ========== RIGHT FRAME: Products ==========
        right_frame = Frame(content_frame, bg="white", bd=1, relief=SOLID, highlightbackground="#dddddd", highlightthickness=1)
        right_frame.grid(row=0, column=1, sticky="nsew")
        right_frame.grid_rowconfigure(4, weight=1)
        right_frame.grid_columnconfigure(0, weight=1)
        
        # Product title
        prod_title = Label(right_frame, text="Manage Products", 
                          font=("Segoe UI", 14, "bold"), 
                          bg=self.primary_color, fg="white", padx=10, pady=5)
        prod_title.grid(row=0, column=0, sticky="ew")
        
        # Low stock info at top right of right frame (floating)
        low_stock_frame = Frame(right_frame, bg="#FEF9E7", bd=1, relief=SOLID, highlightbackground="#F1C40F")
        low_stock_frame.place(relx=1.0, x=-10, y=5, anchor="ne", width=200, height=32)
        
        self.lbl_low_stock_count = Label(low_stock_frame, 
                                        text="Low Stock Items: 0", 
                                        font=("Segoe UI", 10, "bold"), 
                                        bg="#FEF9E7", fg="#B7950B")
        self.lbl_low_stock_count.pack(side=LEFT, padx=8, pady=5)
        
        btn_view_low_stock = Button(low_stock_frame, text="View", 
                                   command=self.show_low_stock_products,
                                   font=("Segoe UI", 8, "bold"), 
                                   bg=self.accent_color, fg="#2C3E50", 
                                   cursor="hand2", width=6, bd=0,
                                   activebackground="#D4AC0D")
        btn_view_low_stock.pack(side=RIGHT, padx=5, pady=3)
        
        # Product Entry Frame (grid with uniform columns)
        prod_entry_frame = Frame(right_frame, bg="white", bd=1, relief=SOLID, padx=10, pady=10)
        prod_entry_frame.grid(row=1, column=0, sticky="ew", padx=5, pady=5)
        
        # Configure columns for uniform entry widths:
        for i in range(4):
            if i in (1, 3):
                prod_entry_frame.grid_columnconfigure(i, weight=1, uniform="prod_entry")
            else:
                prod_entry_frame.grid_columnconfigure(i, weight=0)
        
        # Row 0: Category (left) and empty right placeholder
        lbl_category = Label(prod_entry_frame, text="Category", 
                            font=("Segoe UI", 10), bg="white", fg="#2C3E50")
        lbl_category.grid(row=0, column=0, sticky=W, pady=6, padx=2)
        
        cat_frame = Frame(prod_entry_frame, bg="white")
        cat_frame.grid(row=0, column=1, pady=6, padx=2, sticky="ew")
        cat_frame.grid_columnconfigure(0, weight=1)
        
        self.cat_dropdown = Entry(cat_frame, textvariable=self.var_ProdCat, 
                                 font=("Segoe UI", 10), bg="#F8F9F9", 
                                 fg="#2C3E50", bd=1, relief=SOLID)
        self.cat_dropdown.grid(row=0, column=0, sticky="ew")
        
        self.cat_btn = Button(cat_frame, text="▼", font=("Arial", 8), 
                             width=2, command=self.show_cat_dropdown,
                             bg="#E5E7E9", fg="#2C3E50", bd=1, relief=SOLID)
        self.cat_btn.grid(row=0, column=1, padx=(2,0))
        
        self.cat_listbox = Listbox(prod_entry_frame, font=("Segoe UI", 10), 
                                  height=5, bg="white", fg="#2C3E50",
                                  bd=1, relief=SOLID, selectbackground=self.primary_light)
        self.cat_listbox.bind('<<ListboxSelect>>', self.on_cat_select)
        
        # Placeholder in column 2-3 for row 0
        lbl_empty = Label(prod_entry_frame, text="", bg="white")
        lbl_empty.grid(row=0, column=2, columnspan=2, sticky="ew", pady=6)
        
        # Row 1: Product Name (left) and Barcode (right) - Barcode optional
        lbl_product_name = Label(prod_entry_frame, text="Product Name", 
                                font=("Segoe UI", 10), bg="white", fg="#2C3E50")
        lbl_product_name.grid(row=1, column=0, sticky=W, pady=6, padx=2)
        
        txt_prod_name = Entry(prod_entry_frame, textvariable=self.var_prod_name,
                             font=("Segoe UI", 10), bg="#F8F9F9", 
                             fg="#2C3E50", bd=1, relief=SOLID)
        txt_prod_name.grid(row=1, column=1, pady=6, padx=2, sticky="ew")
        
        lbl_barcode = Label(prod_entry_frame, text="Barcode (Optional)",   # Updated label
                           font=("Segoe UI", 10), bg="white", fg="#2C3E50")
        lbl_barcode.grid(row=1, column=2, sticky=W, pady=6, padx=5)
        
        txt_barcode = Entry(prod_entry_frame, textvariable=self.var_barcode,
                           font=("Segoe UI", 10), bg="#F8F9F9", 
                           fg="#2C3E50", bd=1, relief=SOLID)
        txt_barcode.grid(row=1, column=3, pady=6, padx=2, sticky="ew")
        
        # Row 2: Cost Price (left) and Retail Price (right)
        lbl_cost_price = Label(prod_entry_frame, text="Cost Price", 
                              font=("Segoe UI", 10), bg="white", fg="#2C3E50")
        lbl_cost_price.grid(row=2, column=0, sticky=W, pady=6, padx=2)
        
        cost_price_frame = Frame(prod_entry_frame, bg="white")
        cost_price_frame.grid(row=2, column=1, pady=6, padx=2, sticky="ew")
        cost_price_frame.grid_columnconfigure(1, weight=1)
        
        rs_label_cost = Label(cost_price_frame, text="Rs.", 
                             font=("Segoe UI", 10, "bold"), bg="white", fg=self.primary_dark)
        rs_label_cost.grid(row=0, column=0, padx=(0,3))
        
        self.txt_cost_price = Entry(cost_price_frame, font=("Segoe UI", 10), 
                                   bg="#F8F9F9", fg="#2C3E50", 
                                   justify=RIGHT, bd=1, relief=SOLID)
        self.txt_cost_price.grid(row=0, column=1, sticky="ew")
        
        self.txt_cost_price.bind('<KeyRelease>', self.format_cost_price_with_commas)
        self.txt_cost_price.bind('<FocusIn>', self.on_cost_price_focus_in)
        self.txt_cost_price.bind('<FocusOut>', self.on_cost_price_focus_out)
        self.txt_cost_price.bind('<KeyRelease>', self.calculate_margin)
        
        lbl_price = Label(prod_entry_frame, text="Retail Price", 
                         font=("Segoe UI", 10), bg="white", fg="#2C3E50")
        lbl_price.grid(row=2, column=2, sticky=W, pady=6, padx=5)
        
        price_frame = Frame(prod_entry_frame, bg="white")
        price_frame.grid(row=2, column=3, pady=6, padx=2, sticky="ew")
        price_frame.grid_columnconfigure(1, weight=1)
        
        rs_label = Label(price_frame, text="Rs.", 
                        font=("Segoe UI", 10, "bold"), bg="white", fg=self.success_color)
        rs_label.grid(row=0, column=0, padx=(0,3))
        
        self.txt_price = Entry(price_frame, font=("Segoe UI", 10), 
                              bg="#F8F9F9", fg="#2C3E50", 
                              justify=RIGHT, bd=1, relief=SOLID)
        self.txt_price.grid(row=0, column=1, sticky="ew")
        
        self.txt_price.bind('<KeyRelease>', self.format_price_with_commas)
        self.txt_price.bind('<FocusIn>', self.on_price_focus_in)
        self.txt_price.bind('<FocusOut>', self.on_price_focus_out)
        self.txt_price.bind('<KeyRelease>', self.calculate_margin)
        
        # Row 3: Margin (left) and Quantity (right)
        lbl_margin = Label(prod_entry_frame, text="Margin", 
                          font=("Segoe UI", 10), bg="white", fg="#2C3E50")
        lbl_margin.grid(row=3, column=0, sticky=W, pady=6, padx=2)
        
        margin_frame = Frame(prod_entry_frame, bg="white")
        margin_frame.grid(row=3, column=1, pady=6, padx=2, sticky="ew")
        margin_frame.grid_columnconfigure(1, weight=1)
        
        rs_label_margin = Label(margin_frame, text="Rs.", 
                               font=("Segoe UI", 10, "bold"), bg="white", fg=self.accent_color)
        rs_label_margin.grid(row=0, column=0, padx=(0,3))
        
        self.txt_margin = Entry(margin_frame, textvariable=self.var_margin,
                               font=("Segoe UI", 10), bg="#F2F3F4", 
                               fg="#2C3E50", state="readonly", 
                               bd=1, relief=SOLID, readonlybackground="#F8F9F9")
        self.txt_margin.grid(row=0, column=1, sticky="ew")
        
        lbl_quantity = Label(prod_entry_frame, text="Quantity", 
                            font=("Segoe UI", 10), bg="white", fg="#2C3E50")
        lbl_quantity.grid(row=3, column=2, sticky=W, pady=6, padx=5)
        
        txt_qty = Entry(prod_entry_frame, textvariable=self.var_qty,
                       font=("Segoe UI", 10), bg="#F8F9F9", 
                       fg="#2C3E50", bd=1, relief=SOLID)
        txt_qty.grid(row=3, column=3, pady=6, padx=2, sticky="ew")
        txt_qty.bind('<KeyRelease>', self.check_reorder_status)
        
        # Row 4: Status (left) and Product ID (right)
        lbl_status = Label(prod_entry_frame, text="Status", 
                          font=("Segoe UI", 10), bg="white", fg="#2C3E50")
        lbl_status.grid(row=4, column=0, sticky=W, pady=6, padx=2)
        
        self.cmb_status = ttk.Combobox(prod_entry_frame, textvariable=self.var_status,
                                 values=("Active", "Inactive", "Reorder"), state="readonly",
                                 font=("Segoe UI", 10))
        self.cmb_status.grid(row=4, column=1, pady=6, padx=2, sticky="ew")
        self.cmb_status.set("Active")
        self.cmb_status.bind('<<ComboboxSelected>>', self.on_status_change)
        
        lbl_pid = Label(prod_entry_frame, text="Product ID", 
                       font=("Segoe UI", 10), bg="white", fg="#2C3E50")
        lbl_pid.grid(row=4, column=2, sticky=W, pady=6, padx=5)
        
        txt_pid = Entry(prod_entry_frame, textvariable=self.var_pid,
                       font=("Segoe UI", 10), bg="#F2F3F4", 
                       fg="#2C3E50", state="readonly", 
                       bd=1, relief=SOLID)
        txt_pid.grid(row=4, column=3, pady=6, padx=2, sticky="ew")
        
        # Row 5: Reorder info (spanning columns)
        reorder_info = Label(prod_entry_frame, 
                           text=f"Note: Quantity ≤ {self.reorder_threshold} will auto-set status to 'Reorder'", 
                           font=("Segoe UI", 8, "italic"), 
                           bg="white", fg=self.danger_color)
        reorder_info.grid(row=5, column=0, columnspan=4, sticky=W, pady=(0,3), padx=2)
        
        # Row 6: Product Buttons Frame (uniform width)
        prod_btn_frame = Frame(prod_entry_frame, bg="white")
        prod_btn_frame.grid(row=6, column=0, columnspan=4, pady=12, sticky="ew")
        for i in range(5):
            prod_btn_frame.grid_columnconfigure(i, weight=1, uniform="prod_btn")
        
        # Save - Bright Green
        btn_add_prod = Button(prod_btn_frame, text="SAVE", command=self.add_product,
                            font=("Segoe UI", 11, "bold"),
                            bg=self.btn_save_color, activebackground=self.btn_save_hover,
                            fg="white", cursor="hand2", bd=0)
        btn_add_prod.grid(row=0, column=0, padx=3, pady=3, sticky="ew")
        
        # Edit - Bright Royal Blue
        btn_update_prod = Button(prod_btn_frame, text="EDIT", command=self.update_product,
                               font=("Segoe UI", 11, "bold"),
                               bg=self.btn_update_color, activebackground=self.btn_update_hover,
                               fg="white", cursor="hand2", bd=0)
        btn_update_prod.grid(row=0, column=1, padx=3, pady=3, sticky="ew")
        
        # Delete - Bright Red
        btn_delete_prod = Button(prod_btn_frame, text="DELETE", command=self.delete_product,
                               font=("Segoe UI", 11, "bold"),
                               bg=self.btn_delete_color, activebackground=self.btn_delete_hover,
                               fg="white", cursor="hand2", bd=0)
        btn_delete_prod.grid(row=0, column=2, padx=3, pady=3, sticky="ew")
        
        # Clear - Bright Orange
        btn_clear_prod = Button(prod_btn_frame, text="CLEAR", command=self.clear_product,
                              font=("Segoe UI", 11, "bold"),
                              bg=self.btn_clear_color, activebackground=self.btn_clear_hover,
                              fg="white", cursor="hand2", bd=0)
        btn_clear_prod.grid(row=0, column=3, padx=3, pady=3, sticky="ew")
        
        # (Placeholder for a future button or empty)
        btn_placeholder = Label(prod_btn_frame, text="", bg="white")
        btn_placeholder.grid(row=0, column=4, padx=3, pady=3, sticky="ew")
        
        # Search Frame
        search_frame = LabelFrame(right_frame, text="Search Product", 
                                 font=("Segoe UI", 10, "bold"), 
                                 bd=1, relief=SOLID, bg="white", fg=self.primary_color)
        search_frame.grid(row=2, column=0, sticky="ew", padx=5, pady=5)
        search_frame.grid_columnconfigure(1, weight=1)
        
        cmb_search = ttk.Combobox(search_frame, textvariable=self.var_searchby,
                                 values=("Select", "Category", "Name", "Status", "Low Stock (≤5)"),
                                 state="readonly", font=("Segoe UI", 10), width=12)
        cmb_search.grid(row=0, column=0, padx=5, pady=5)
        cmb_search.set("Select")
        
        txt_search = Entry(search_frame, textvariable=self.var_searchtxt,
                          font=("Segoe UI", 10), bg="#F8F9F9", 
                          fg="#2C3E50", bd=1, relief=SOLID)
        txt_search.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        
        btn_search = Button(search_frame, text="SEARCH", command=self.search_product,
                           font=("Segoe UI", 9, "bold"), bg="#5DADE2", 
                           fg="white", cursor="hand2", bd=0, padx=10)
        btn_search.grid(row=0, column=2, padx=2, pady=5)
        
        btn_showall = Button(search_frame, text="SHOW ALL", command=self.show_products,
                            font=("Segoe UI", 9, "bold"), bg=self.primary_color, 
                            fg="white", cursor="hand2", bd=0, padx=10)
        btn_showall.grid(row=0, column=3, padx=2, pady=5)
        
        btn_import = Button(search_frame, text="IMPORT", command=self.import_excel_dialog,
                           font=("Segoe UI", 9, "bold"), bg=self.btn_save_color, 
                           fg="white", cursor="hand2", bd=0, padx=10)
        btn_import.grid(row=0, column=4, padx=2, pady=5)
        
        btn_print = Button(search_frame, text="PRINT PDF", command=self.print_all_products,
                          font=("Segoe UI", 9, "bold"),
                          bg=self.btn_pdf_color, activebackground=self.btn_pdf_hover,
                          fg="white", cursor="hand2", bd=0, padx=10)
        btn_print.grid(row=0, column=5, padx=2, pady=5)
        
        # Products Table Frame
        prod_table_frame = Frame(right_frame, bd=1, relief=SOLID, bg="white")
        prod_table_frame.grid(row=4, column=0, sticky="nsew", padx=5, pady=5)
        prod_table_frame.grid_rowconfigure(0, weight=1)
        prod_table_frame.grid_columnconfigure(0, weight=1)
        
        scrolly_prod = Scrollbar(prod_table_frame, orient=VERTICAL)
        scrollx_prod = Scrollbar(prod_table_frame, orient=HORIZONTAL)
        
        # Added "Barcode" column
        self.Product_Table = ttk.Treeview(prod_table_frame, 
                                         columns=("pid", "Category", "Name", "Barcode", "Cost Price", "Retail Price", "Margin", "Qty", "Status"),
                                         yscrollcommand=scrolly_prod.set,
                                         xscrollcommand=scrollx_prod.set)
        
        scrollx_prod.pack(side=BOTTOM, fill=X)
        scrolly_prod.pack(side=RIGHT, fill=Y)
        scrollx_prod.config(command=self.Product_Table.xview)
        scrolly_prod.config(command=self.Product_Table.yview)
        
        # Configure headings
        self.Product_Table.heading("pid", text="ID")
        self.Product_Table.heading("Category", text="Category")
        self.Product_Table.heading("Name", text="Product Name")
        self.Product_Table.heading("Barcode", text="Barcode")
        self.Product_Table.heading("Cost Price", text="Cost Price")
        self.Product_Table.heading("Retail Price", text="Retail Price")
        self.Product_Table.heading("Margin", text="Margin")
        self.Product_Table.heading("Qty", text="Qty")
        self.Product_Table.heading("Status", text="Status")
        
        self.Product_Table["show"] = "headings"
        
        self.Product_Table.column("pid", width=40, anchor=CENTER)
        self.Product_Table.column("Category", width=100)
        self.Product_Table.column("Name", width=120)
        self.Product_Table.column("Barcode", width=100, anchor=CENTER)
        self.Product_Table.column("Cost Price", width=90, anchor=E)
        self.Product_Table.column("Retail Price", width=90, anchor=E)
        self.Product_Table.column("Margin", width=90, anchor=E)
        self.Product_Table.column("Qty", width=50, anchor=CENTER)
        self.Product_Table.column("Status", width=80, anchor=CENTER)
        
        self.Product_Table.tag_configure('reorder', background='#FEF9E7', foreground='#B7950B')
        self.Product_Table.tag_configure('active', background='#D5F5E3', foreground='#145A32')
        self.Product_Table.tag_configure('inactive', background='#FADBD8', foreground='#943126')
        
        self.Product_Table.pack(fill=BOTH, expand=True)
        self.Product_Table.bind("<ButtonRelease-1>", self.get_product_data)
        
        # Initialize data
        self.show_categories()
        self.fetch_categories()
        self.generate_pid()
        self.show_products()
        
        self.Category_Table.bind("<<TreeviewSelect>>", self.on_category_select)
        
        self.ensure_database_columns()
        self.update_low_stock_count()
    
    # ========== DATE & TIME METHODS ==========
    def update_datetime(self):
        now = datetime.now()
        date_str = now.strftime("%A, %B %d, %Y")
        time_str = now.strftime("%I:%M:%S %p")
        self.lbl_datetime.config(text=f"{date_str}\n{time_str}")
        self.root.after(1000, self.update_datetime)
    
    # ========== DATABASE INITIALIZATION ==========
    def ensure_database_columns(self):
        """Ensure all required columns exist in the database tables"""
        con = sqlite3.connect(database=r'Possystem.db')
        cur = con.cursor()
        try:
            # Check product table columns
            cur.execute("PRAGMA table_info(product)")
            columns = [col[1] for col in cur.fetchall()]
            
            # Add barcode column if missing
            if 'barcode' not in columns:
                try:
                    cur.execute("ALTER TABLE product ADD COLUMN barcode TEXT")
                except Exception as e:
                    print(f"Error adding barcode: {e}")
            
            if 'cost_price' not in columns:
                try:
                    cur.execute("ALTER TABLE product ADD COLUMN cost_price REAL DEFAULT 0")
                except Exception as e:
                    print(f"Error adding cost_price: {e}")
            
            if 'retailprice' not in columns:
                try:
                    cur.execute("ALTER TABLE product ADD COLUMN retailprice REAL DEFAULT 0")
                except Exception as e:
                    print(f"Error adding retailprice: {e}")
            
            if 'margin' not in columns:
                try:
                    cur.execute("ALTER TABLE product ADD COLUMN margin REAL DEFAULT 0")
                except Exception as e:
                    print(f"Error adding margin: {e}")
            
            con.commit()
        except Exception as ex:
            print(f"Error ensuring database columns: {ex}")
        finally:
            con.close()
    
    # ========== LOW STOCK METHODS ==========
    def update_low_stock_count(self):
        con = sqlite3.connect(database=r'Possystem.db')
        cur = con.cursor()
        try:
            cur.execute("SELECT COUNT(*) FROM product WHERE Quantity <= ? AND (Status != 'Inactive' OR Status IS NULL)", 
                       (self.reorder_threshold,))
            count = cur.fetchone()[0]
            self.lbl_low_stock_count.config(text=f"Low Stock Items: {count}")
            self.root.after(10000, self.update_low_stock_count)
        except Exception as ex:
            print(f"Error updating low stock count: {ex}")
        finally:
            con.close()
    
    def show_low_stock_products(self):
        self.var_searchby.set("Low Stock (≤5)")
        self.var_searchtxt.set("")
        self.search_product()
    
    # ========== REORDER STATUS ==========
    def determine_status(self, quantity, current_status=None):
        try:
            qty = int(quantity) if quantity else 0
        except:
            qty = 0
        
        if qty <= self.reorder_threshold:
            return "Reorder"
        elif current_status and current_status not in ["Reorder", "Select"]:
            return current_status
        else:
            return "Active"
    
    def check_reorder_status(self, event=None):
        try:
            qty_text = self.var_qty.get()
            if qty_text:
                qty = int(qty_text)
                if qty <= self.reorder_threshold:
                    self.var_status.set("Reorder")
                    self.cmb_status.config(foreground="#B7950B", background="#FEF9E7")
                else:
                    if self.var_status.get() == "Reorder":
                        self.var_status.set("Active")
                        self.cmb_status.config(foreground="black", background="white")
        except ValueError:
            pass
    
    def on_status_change(self, event=None):
        pass
    
    # ========== CATEGORY DROPDOWN ==========
    def show_cat_dropdown(self, event=None):
        parent = self.cat_listbox.master
        x = self.cat_dropdown.winfo_rootx() - parent.winfo_rootx()
        y = self.cat_dropdown.winfo_rooty() - parent.winfo_rooty() + self.cat_dropdown.winfo_height()
        
        self.cat_listbox.place(x=x, y=y, width=200, height=100)
        self.cat_listbox.delete(0, END)
        for cat in self.cat_list:
            if cat != "Select":
                self.cat_listbox.insert(END, cat)
        self.cat_listbox.lift()
        self.cat_listbox.focus_set()

    def filter_categories(self, event=None):
        search_text = self.var_ProdCat.get().lower()
        self.cat_listbox.delete(0, END)
        for cat in self.cat_list:
            if cat != "Select" and search_text in cat.lower():
                self.cat_listbox.insert(END, cat)
        parent = self.cat_listbox.master
        x = self.cat_dropdown.winfo_rootx() - parent.winfo_rootx()
        y = self.cat_dropdown.winfo_rooty() - parent.winfo_rooty() + self.cat_dropdown.winfo_height()
        if self.cat_listbox.size() > 0 and search_text:
            self.cat_listbox.place(x=x, y=y, width=200, height=100)
        else:
            self.cat_listbox.place_forget()
    
    def on_cat_select(self, event=None):
        selection = self.cat_listbox.curselection()
        if selection:
            selected_cat = self.cat_listbox.get(selection[0])
            self.var_ProdCat.set(selected_cat)
            self.cat_listbox.place_forget()
    
    # ========== CATEGORY METHODS ==========
    def get_next_category_id(self):
        con = sqlite3.connect(database=r'Possystem.db')
        cur = con.cursor()
        try:
            cur.execute("SELECT CID FROM Category ORDER BY CID")
            rows = cur.fetchall()
            if not rows:
                return 1
            ids = [row[0] for row in rows]
            for i in range(1, len(ids) + 2):
                if i not in ids:
                    return i
            return len(ids) + 1
        finally:
            con.close()
    
    def add_category(self):
        con = sqlite3.connect(database=r'Possystem.db')
        cur = con.cursor()
        try:
            if self.var_CatName.get() == "":
                messagebox.showerror("Error", "Category Name is required", parent=self.root)
            else:
                cur.execute("SELECT * FROM Category WHERE Name=?", (self.var_CatName.get(),))
                row = cur.fetchone()
                if row is not None:
                    messagebox.showerror("Error", "Category already exists", parent=self.root)
                else:
                    next_id = self.get_next_category_id()
                    cur.execute("INSERT INTO Category (CID, Name) VALUES(?, ?)", 
                               (next_id, self.var_CatName.get()))
                    con.commit()
                    messagebox.showinfo("Success", "Category added successfully", parent=self.root)
                    self.show_categories()
                    self.fetch_categories()
                    self.var_CatName.set("")
        except Exception as ex:
            messagebox.showerror("Error", f"Error due to: {str(ex)}", parent=self.root)
        finally:
            con.close()
    
    def show_categories(self):
        con = sqlite3.connect(database=r'Possystem.db')
        cur = con.cursor()
        try:
            cur.execute("SELECT * FROM Category ORDER BY CID")
            rows = cur.fetchall()
            self.Category_Table.delete(*self.Category_Table.get_children())
            for row in rows:
                self.Category_Table.insert('', END, values=row)
        except Exception as ex:
            messagebox.showerror("Error", f"Error loading categories: {str(ex)}", parent=self.root)
        finally:
            con.close()
    
    def get_category_data(self, ev):
        try:
            selected_item = self.Category_Table.selection()[0]
            row = self.Category_Table.item(selected_item, 'values')
            if row:
                self.var_CatID.set(row[0])
                self.var_CatName.set(row[1])
        except IndexError:
            pass
    
    def delete_category(self):
        con = sqlite3.connect(database='Possystem.db')
        cur = con.cursor()
        try:
            if not self.var_CatID.get():
                messagebox.showerror("Error", "Please select a category to delete", parent=self.root)
            else:
                cur.execute("SELECT COUNT(*) FROM product WHERE Category=?", 
                           (self.var_CatName.get(),))
                product_count = cur.fetchone()[0]
                if product_count > 0:
                    messagebox.showerror("Error", 
                                       f"Cannot delete category. It has {product_count} product(s) associated with it.\nPlease delete or reassign the products first.", 
                                       parent=self.root)
                    return
                op = messagebox.askyesno("Confirm", "Do you really want to delete this category?", parent=self.root)
                if op:
                    delete_id = int(self.var_CatID.get())
                    cur.execute("DELETE FROM Category WHERE CID=?", (delete_id,))
                    cur.execute("SELECT CID FROM Category WHERE CID > ? ORDER BY CID", (delete_id,))
                    higher_ids = cur.fetchall()
                    for old_id in higher_ids:
                        new_id = old_id[0] - 1
                        cur.execute("UPDATE Category SET CID = ? WHERE CID = ?", (new_id, old_id[0]))
                    con.commit()
                    messagebox.showinfo("Delete", "Category deleted successfully", parent=self.root)
                    self.show_categories()
                    self.fetch_categories()
                    self.var_CatID.set("")
                    self.var_CatName.set("")
        except Exception as ex:
            messagebox.showerror("Error", f"Error due to: {str(ex)}", parent=self.root)
        finally:
            con.close()
    
    def clear_category(self):
        self.var_CatID.set("")
        self.var_CatName.set("")
        self.Category_Table.selection_remove(self.Category_Table.selection())
    
    def on_category_select(self, event):
        try:
            selected_item = self.Category_Table.selection()[0]
            row = self.Category_Table.item(selected_item, 'values')
            if row:
                category_name = row[1]
                self.var_ProdCat.set(category_name)
                self.show_products_by_category(category_name)
        except IndexError:
            pass
    
    # ========== MARGIN CALCULATION ==========
    def calculate_margin(self, event=None):
        try:
            cost_price = self.get_cost_price_value()
            retail_price = self.get_price_value()
            if cost_price is not None and retail_price is not None:
                margin = retail_price - cost_price
                self.var_margin.set(f"{margin:,.2f}")
            else:
                self.var_margin.set("")
        except Exception as e:
            print(f"Error calculating margin: {e}")
            self.var_margin.set("")
    
    # ========== PRICE FORMATTING ==========
    def format_price_with_commas(self, event=None):
        try:
            current_text = self.txt_price.get()
            cursor_pos = self.txt_price.index(INSERT)
            clean_text = ''
            has_decimal = False
            for char in current_text:
                if char.isdigit():
                    clean_text += char
                elif char == '.' and not has_decimal:
                    clean_text += char
                    has_decimal = True
                elif char == ',':
                    continue
            self.price_without_format = clean_text
            if clean_text:
                if '.' in clean_text:
                    integer_part, decimal_part = clean_text.split('.')
                else:
                    integer_part = clean_text
                    decimal_part = ''
                formatted_integer = ''
                for i, digit in enumerate(reversed(integer_part)):
                    if i > 0 and i % 3 == 0:
                        formatted_integer = ',' + formatted_integer
                    formatted_integer = digit + formatted_integer
                formatted_price = formatted_integer
                if decimal_part:
                    formatted_price += '.' + decimal_part
                self.txt_price.delete(0, END)
                self.txt_price.insert(0, formatted_price)
                try:
                    self.txt_price.icursor(cursor_pos)
                except:
                    pass
        except Exception as e:
            print(f"Error formatting price: {e}")
    
    def on_price_focus_in(self, event=None):
        self.price_without_format = self.txt_price.get().replace(',', '')
    
    def on_price_focus_out(self, event=None):
        self.format_price_with_commas()
        self.calculate_margin()
    
    def get_price_value(self):
        try:
            price_text = self.txt_price.get().replace(',', '')
            if not price_text:
                return None
            return float(price_text)
        except ValueError:
            return None
    
    def format_cost_price_with_commas(self, event=None):
        try:
            current_text = self.txt_cost_price.get()
            cursor_pos = self.txt_cost_price.index(INSERT)
            clean_text = ''
            has_decimal = False
            for char in current_text:
                if char.isdigit():
                    clean_text += char
                elif char == '.' and not has_decimal:
                    clean_text += char
                    has_decimal = True
                elif char == ',':
                    continue
            self.cost_price_without_format = clean_text
            if clean_text:
                if '.' in clean_text:
                    integer_part, decimal_part = clean_text.split('.')
                else:
                    integer_part = clean_text
                    decimal_part = ''
                formatted_integer = ''
                for i, digit in enumerate(reversed(integer_part)):
                    if i > 0 and i % 3 == 0:
                        formatted_integer = ',' + formatted_integer
                    formatted_integer = digit + formatted_integer
                formatted_price = formatted_integer
                if decimal_part:
                    formatted_price += '.' + decimal_part
                self.txt_cost_price.delete(0, END)
                self.txt_cost_price.insert(0, formatted_price)
                try:
                    self.txt_cost_price.icursor(cursor_pos)
                except:
                    pass
        except Exception as e:
            print(f"Error formatting cost price: {e}")
    
    def on_cost_price_focus_in(self, event=None):
        self.cost_price_without_format = self.txt_cost_price.get().replace(',', '')
    
    def on_cost_price_focus_out(self, event=None):
        self.format_cost_price_with_commas()
        self.calculate_margin()
    
    def get_cost_price_value(self):
        try:
            price_text = self.txt_cost_price.get().replace(',', '')
            if not price_text:
                return None
            return float(price_text)
        except ValueError:
            return None
    
    # ========== CATEGORY FETCHING ==========
    def fetch_categories(self):
        con = sqlite3.connect(database=r'Possystem.db')
        cur = con.cursor()
        try:
            self.cat_list.clear()
            try:
                cur.execute("PRAGMA table_info(Category)")
                columns = [col[1] for col in cur.fetchall()]
                if 'Status' in columns:
                    cur.execute("SELECT Name FROM Category WHERE Status='Active' ORDER BY Name")
                else:
                    cur.execute("SELECT Name FROM Category ORDER BY Name")
            except:
                cur.execute("SELECT Name FROM Category ORDER BY Name")
            categories = cur.fetchall()
            self.cat_list = ["Select"] + [cat[0] for cat in categories]
        except Exception as ex:
            messagebox.showerror("Error", f"Error fetching categories: {str(ex)}", parent=self.root)
            self.cat_list = ["Select"]
        finally:
            con.close()
    
    def generate_pid(self):
        try:
            con = sqlite3.connect(database=r'Possystem.db')
            cur = con.cursor()
            cur.execute("SELECT pid FROM product ORDER BY pid")
            rows = cur.fetchall()
            if not rows:
                self.var_pid.set("001")
                return
            existing_ids = []
            for row in rows:
                try:
                    pid_value = int(row[0]) if row[0] is not None else 0
                    existing_ids.append(pid_value)
                except (ValueError, TypeError):
                    continue
            if not existing_ids:
                self.var_pid.set("001")
                return
            for i in range(1, len(existing_ids) + 2):
                if i not in existing_ids:
                    self.var_pid.set(f"{i:03d}")
                    return
            self.var_pid.set(f"{max(existing_ids) + 1:03d}")
        except Exception as ex:
            messagebox.showerror("Error", f"Error generating PID: {str(ex)}", parent=self.root)
            self.var_pid.set("001")
        finally:
            con.close()
    
    # ========== EXCEL IMPORT ==========
    def import_excel_dialog(self):
        instructions = f"""EXCEL IMPORT INSTRUCTIONS:

Your Excel file MUST have the following 6 columns (exact column names):
1. Category      - Product category (must exist in system or will be created)
2. Product_Name  - Name of the product
3. Barcode       - Product barcode (optional)                           # UPDATED
4. Cost_Price    - Product cost price
5. Retail_Price  - Product retail price  
6. Quantity      - Quantity in stock (whole numbers)

OPTIONAL:
- Add 'Status' column with values 'Active', 'Inactive', or 'Reorder' (default: Auto-calculated based on quantity)
- The system will auto-generate Product IDs and calculate Margin automatically

⚠️ IMPORTANT: Products with Quantity ≤ {self.reorder_threshold} will automatically be set to 'Reorder' status in the Status column

Example Excel format:
┌──────────┬───────────────┬─────────────┬────────────┬──────────────┬─────────┐
│ Category │ Product_Name  │   Barcode   │ Cost_Price │ Retail_Price │ Quantity│
├──────────┼───────────────┼─────────────┼────────────┼──────────────┼─────────┤
│Electronics│ Laptop Dell   │ 123456789012│   75000    │    85000     │   10    │
│Furniture │ Office Chair  │ 234567890123│   10000    │    12500     │   15    │
│Stationery│ A4 Paper Pack │ 345678901234│    350     │     450      │   50    │
│Groceries │ Sugar 1kg     │ (empty)     │    100     │     120      │    **3** │ → Barcode can be blank, 'Reorder' in Status
└──────────┴───────────────┴─────────────┴────────────┴──────────────┴─────────┘

Note: The file can be .xlsx or .xls format."""
        
        messagebox.showinfo("Excel Import Instructions", instructions, parent=self.root)
        
        response = messagebox.askyesno("Import Excel", 
                                      "Do you have an Excel file ready for import?\n\nClick YES to select your Excel file.\nClick NO to download a template.", 
                                      parent=self.root)
        
        if response:
            self.select_excel_file()
        else:
            self.generate_excel_template()
    
    def generate_excel_template(self):
        try:
            sample_data = {
                'Category': ['Electronics', 'Furniture', 'Stationery', 'Groceries'],
                'Product_Name': ['Laptop Dell', 'Office Chair', 'A4 Paper Pack', 'Sugar 1kg'],
                'Barcode': ['123456789012', '234567890123', '', ''],   # Show optional
                'Cost_Price': [75000, 10000, 350, 100],
                'Retail_Price': [85000, 12500, 450, 120],
                'Quantity': [10, 15, 50, 3],
                'Status': ['Active', 'Active', 'Active', '']
            }
            
            df = pd.DataFrame(sample_data)
            
            file_path = filedialog.asksaveasfilename(
                parent=self.root,
                title="Save Excel Template",
                defaultextension=".xlsx",
                initialfile="Inventory_Import_Template.xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("Excel 97-2003", "*.xls"), ("All files", "*.*")]
            )
            
            if file_path:
                df.to_excel(file_path, index=False)
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                for cell in ws[1]:
                    cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
                    cell.fill = openpyxl.styles.PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
                    cell.alignment = openpyxl.styles.Alignment(horizontal="center")
                ws['H1'] = 'Note:'
                ws['H2'] = f'Products with Quantity ≤ {self.reorder_threshold} will auto-set to "Reorder" in Status column'
                ws['H2'].font = openpyxl.styles.Font(italic=True, color="FF0000")
                ws['H3'] = 'Barcode is optional; leave blank if not used.'
                ws['H3'].font = openpyxl.styles.Font(italic=True, color="9B59B6")
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                wb.save(file_path)
                
                open_file = messagebox.askyesno("Template Created", 
                                               f"Excel template saved at:\n{file_path}\n\nDo you want to open it now to edit?", 
                                               parent=self.root)
                if open_file:
                    os.startfile(file_path)
                    
                import_now = messagebox.askyesno("Import File", 
                                                "Template created! Do you want to import an Excel file now?", 
                                                parent=self.root)
                if import_now:
                    self.select_excel_file()
        except Exception as ex:
            messagebox.showerror("Error", f"Error generating template: {str(ex)}", parent=self.root)
    
    def select_excel_file(self):
        file_path = filedialog.askopenfilename(
            parent=self.root,
            title="Select Excel File to Import",
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("Excel 2007+", "*.xlsx"),
                ("Excel 97-2003", "*.xls"),
                ("All files", "*.*")
            ]
        )
        if file_path:
            self.preview_excel_data(file_path)
    
    def preview_excel_data(self, file_path):
        try:
            df = pd.read_excel(file_path)
            # Required columns (Barcode is now optional but still in list for mapping)
            required_columns = ['Category', 'Product_Name', 'Barcode', 'Cost_Price', 'Retail_Price', 'Quantity']
            df_columns_lower = [str(col).strip().lower() for col in df.columns]
            required_columns_lower = [col.lower() for col in required_columns]
            missing_columns = []
            for req_col, req_col_lower in zip(required_columns, required_columns_lower):
                if req_col_lower not in df_columns_lower:
                    missing_columns.append(req_col)
            if missing_columns:
                messagebox.showerror("Error", 
                                   f"Missing required columns:\n{', '.join(missing_columns)}\n\nPlease check your Excel file format.",
                                   parent=self.root)
                return
            column_mapping = {}
            for df_col in df.columns:
                df_col_lower = str(df_col).strip().lower()
                for req_col, req_col_lower in zip(required_columns, required_columns_lower):
                    if df_col_lower == req_col_lower:
                        column_mapping[req_col] = df_col
                        break
            if len(column_mapping) != len(required_columns):
                messagebox.showerror("Error", 
                                   "Could not map all required columns. Please check column names.",
                                   parent=self.root)
                return
            
            preview_window = Toplevel(self.root)
            preview_window.title("Preview Import Data")
            preview_window.geometry("1100x550+150+100")
            preview_window.config(bg="white")
            preview_window.transient(self.root)
            preview_window.grab_set()
            
            title_label = Label(preview_window, text="PREVIEW IMPORT DATA", 
                              font=("Segoe UI", 14, "bold"), 
                              bg=self.primary_color, fg="white")
            title_label.pack(fill=X, pady=(0, 5))
            
            file_info = Frame(preview_window, bg="white")
            file_info.pack(fill=X, padx=15, pady=5)
            Label(file_info, text=f"File: {os.path.basename(file_path)}", 
                 font=("Segoe UI", 9), bg="white", fg="#2c3e50").pack(anchor=W)
            Label(file_info, text=f"Rows to import: {len(df)}", 
                 font=("Segoe UI", 9), bg="white", fg="#2c3e50").pack(anchor=W)
            Label(file_info, text=f"⚠️ Products with Quantity ≤ {self.reorder_threshold} will be set to 'Reorder' in the Status column", 
                 font=("Segoe UI", 8, "bold"), bg="white", fg="#dc3545").pack(anchor=W, pady=2)
            Label(file_info, text="Barcode is optional; blank cells are allowed.", 
                 font=("Segoe UI", 8), bg="white", fg="#9B59B6").pack(anchor=W)
            
            table_frame = Frame(preview_window, bg="white")
            table_frame.pack(fill=BOTH, expand=True, padx=15, pady=5)
            columns = ['Row'] + required_columns + ['Margin', 'Status']
            preview_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=12)
            style = ttk.Style()
            style.configure("Preview.Treeview", rowheight=22)
            preview_tree.configure(style="Preview.Treeview")
            preview_tree.heading('Row', text='#')
            preview_tree.column('Row', width=35, anchor=CENTER)
            for col in required_columns:
                preview_tree.heading(col, text=col)
                preview_tree.column(col, width=110)
            preview_tree.heading('Margin', text='Margin')
            preview_tree.column('Margin', width=100, anchor=E)
            preview_tree.heading('Status', text='Status')
            preview_tree.column('Status', width=90, anchor=CENTER)
            
            vsb = Scrollbar(table_frame, orient="vertical", command=preview_tree.yview)
            hsb = Scrollbar(table_frame, orient="horizontal", command=preview_tree.xview)
            preview_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            preview_tree.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")
            hsb.grid(row=1, column=0, sticky="ew")
            table_frame.grid_rowconfigure(0, weight=1)
            table_frame.grid_columnconfigure(0, weight=1)
            preview_tree.tag_configure('reorder_preview', background='#fff3cd')
            
            preview_data = df.head(30)
            for idx, row in preview_data.iterrows():
                values = [idx + 1]
                tags = []
                for col in required_columns:
                    value = row[column_mapping[col]]
                    if col in ['Cost_Price', 'Retail_Price']:
                        try:
                            value = f"Rs. {float(value):,.2f}"
                        except:
                            value = str(value)
                    # For barcode, just show as string
                    values.append(value)
                try:
                    cost = float(row[column_mapping['Cost_Price']])
                    retail = float(row[column_mapping['Retail_Price']])
                    margin = retail - cost
                    margin_formatted = f"Rs. {margin:,.2f}"
                except:
                    margin_formatted = "Rs. 0.00"
                values.append(margin_formatted)
                try:
                    qty = int(row[column_mapping['Quantity']])
                    if qty <= self.reorder_threshold:
                        status_val = "Reorder"
                        tags.append('reorder_preview')
                    else:
                        if 'Status' in df.columns and pd.notna(row['Status']):
                            status_val = str(row['Status']).strip()
                            if status_val not in ['Active', 'Inactive', 'Reorder']:
                                status_val = 'Active'
                        else:
                            status_val = 'Active'
                except:
                    status_val = 'Active'
                values.append(status_val)
                preview_tree.insert("", END, values=values, tags=tags)
            
            button_frame = Frame(preview_window, bg="white")
            button_frame.pack(fill=X, padx=15, pady=8)
            import_btn = Button(button_frame, text="IMPORT DATA", 
                              command=lambda: self.import_excel_data(file_path, preview_window),
                              font=("Segoe UI", 9, "bold"), bg=self.success_color, 
                              fg="white", cursor="hand2", width=12)
            import_btn.pack(side=LEFT, padx=5)
            cancel_btn = Button(button_frame, text="CANCEL", 
                              command=preview_window.destroy,
                              font=("Segoe UI", 9, "bold"), bg=self.danger_color, 
                              fg="white", cursor="hand2", width=12)
            cancel_btn.pack(side=LEFT, padx=5)
            if len(df) > 30:
                Label(button_frame, text=f"Showing first 30 of {len(df)} rows", 
                     font=("Segoe UI", 8), bg="white", fg="#6c757d").pack(side=LEFT, padx=15)
        except Exception as ex:
            messagebox.showerror("Error", f"Error reading Excel file:\n{str(ex)}", parent=self.root)
    
    def import_excel_data(self, file_path, preview_window):
        try:
            df = pd.read_excel(file_path)
            required_columns = ['Category', 'Product_Name', 'Barcode', 'Cost_Price', 'Retail_Price', 'Quantity']
            column_mapping = {}
            for df_col in df.columns:
                df_col_lower = str(df_col).strip().lower()
                for req_col in required_columns:
                    if df_col_lower == req_col.lower():
                        column_mapping[req_col] = df_col
                        break
            con = sqlite3.connect(database=r'Possystem.db')
            cur = con.cursor()
            self.ensure_database_columns()
            stats = {
                'total': len(df),
                'added': 0,
                'updated': 0,
                'skipped': 0,
                'categories_added': 0,
                'reorder_count': 0,
                'errors': []
            }
            cur.execute("SELECT Name FROM product")
            existing_products = [row[0].lower() for row in cur.fetchall()]
            cur.execute("SELECT Name FROM Category")
            existing_categories = [row[0] for row in cur.fetchall()]
            # For barcode uniqueness: get all non-empty barcodes
            cur.execute("SELECT barcode FROM product WHERE barcode IS NOT NULL AND barcode != ''")
            existing_barcodes = [row[0] for row in cur.fetchall()]

            for idx, row in df.iterrows():
                try:
                    category = str(row[column_mapping['Category']]).strip()
                    product_name = str(row[column_mapping['Product_Name']]).strip()
                    barcode = str(row[column_mapping['Barcode']]).strip()    # can be empty
                    if product_name.lower() in existing_products:
                        stats['skipped'] += 1
                        stats['errors'].append(f"Row {idx+1}: Product '{product_name}' already exists")
                        continue
                    # Check barcode uniqueness only if it's not empty
                    if barcode and barcode in existing_barcodes:
                        stats['errors'].append(f"Row {idx+1}: Barcode '{barcode}' already exists for another product")
                        stats['skipped'] += 1
                        continue
                    try:
                        cost_price = float(row[column_mapping['Cost_Price']])
                    except:
                        stats['errors'].append(f"Row {idx+1}: Invalid cost price format")
                        stats['skipped'] += 1
                        continue
                    try:
                        retail_price = float(row[column_mapping['Retail_Price']])
                    except:
                        stats['errors'].append(f"Row {idx+1}: Invalid retail price format")
                        stats['skipped'] += 1
                        continue
                    try:
                        quantity = int(row[column_mapping['Quantity']])
                    except:
                        stats['errors'].append(f"Row {idx+1}: Invalid quantity format")
                        stats['skipped'] += 1
                        continue
                    margin = retail_price - cost_price
                    if quantity <= self.reorder_threshold:
                        status = 'Reorder'
                        stats['reorder_count'] += 1
                    else:
                        if 'Status' in df.columns and pd.notna(row['Status']):
                            status = str(row['Status']).strip()
                            if status not in ['Active', 'Inactive', 'Reorder']:
                                status = 'Active'
                        else:
                            status = 'Active'
                    if category not in existing_categories:
                        cur.execute("SELECT CID FROM Category ORDER BY CID")
                        cat_rows = cur.fetchall()
                        cat_ids = [row[0] for row in cat_rows]
                        next_cat_id = 1
                        for i in range(1, len(cat_ids) + 2):
                            if i not in cat_ids:
                                next_cat_id = i
                                break
                        cur.execute("INSERT INTO Category (CID, Name) VALUES(?, ?)", 
                                   (next_cat_id, category))
                        existing_categories.append(category)
                        stats['categories_added'] += 1
                    cur.execute("SELECT pid FROM product ORDER BY pid")
                    prod_rows = cur.fetchall()
                    prod_ids = [int(row[0]) for row in prod_rows if row[0] is not None]
                    next_pid = 1
                    for i in range(1, len(prod_ids) + 2):
                        if i not in prod_ids:
                            next_pid = i
                            break
                    # Insert with barcode (could be empty)
                    cur.execute("""
                        INSERT INTO product (pid, Category, Name, barcode, cost_price, retailprice, margin, Quantity, Status) 
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (next_pid, category, product_name, barcode, cost_price, retail_price, margin, quantity, status))
                    stats['added'] += 1
                    existing_products.append(product_name.lower())
                    if barcode:
                        existing_barcodes.append(barcode)
                except Exception as row_ex:
                    stats['errors'].append(f"Row {idx+1}: {str(row_ex)}")
                    stats['skipped'] += 1
            con.commit()
            con.close()
            preview_window.destroy()
            self.show_import_summary(stats)
            self.show_categories()
            self.fetch_categories()
            self.show_products()
            self.generate_pid()
            self.update_low_stock_count()
        except Exception as ex:
            messagebox.showerror("Error", f"Error importing data:\n{str(ex)}", parent=self.root)
    
    def show_import_summary(self, stats):
        summary_text = f"""
IMPORT SUMMARY:
──────────────
Total rows in file: {stats['total']}
Products added: {stats['added']}
Categories added: {stats['categories_added']}
Products skipped: {stats['skipped']}
Products set to 'Reorder' in Status column: {stats.get('reorder_count', 0)}
"""
        if stats['errors']:
            summary_text += f"\nERRORS ({len(stats['errors'])}):\n"
            for i, error in enumerate(stats['errors'][:10], 1):
                summary_text += f"{i}. {error}\n"
            if len(stats['errors']) > 10:
                summary_text += f"... and {len(stats['errors']) - 10} more errors\n"
        summary_window = Toplevel(self.root)
        summary_window.title("Import Summary")
        summary_window.geometry("550x350+300+200")
        summary_window.config(bg="white")
        summary_window.transient(self.root)
        title_label = Label(summary_window, text="IMPORT COMPLETE", 
                          font=("Segoe UI", 12, "bold"), 
                          bg=self.success_color, fg="white")
        title_label.pack(fill=X, pady=(0, 8))
        text_frame = Frame(summary_window, bg="white")
        text_frame.pack(fill=BOTH, expand=True, padx=15, pady=8)
        summary_text_widget = Text(text_frame, wrap=WORD, font=("Segoe UI", 9), 
                                 bg="#f8f9fa", fg="#2c3e50", height=12)
        scrollbar = Scrollbar(text_frame, command=summary_text_widget.yview)
        summary_text_widget.configure(yscrollcommand=scrollbar.set)
        summary_text_widget.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar.pack(side=RIGHT, fill=Y)
        summary_text_widget.insert(END, summary_text)
        summary_text_widget.config(state=DISABLED)
        button_frame = Frame(summary_window, bg="white")
        button_frame.pack(fill=X, padx=15, pady=8)
        ok_btn = Button(button_frame, text="OK", 
                       command=summary_window.destroy,
                       font=("Segoe UI", 9, "bold"), bg=self.primary_color, 
                       fg="white", cursor="hand2", width=12)
        ok_btn.pack()
    
    # ========== PRODUCT METHODS ==========
    def add_product(self):
        if self.var_ProdCat.get() == "" or self.var_ProdCat.get() == "Select":
            messagebox.showerror("Error", "Category is required", parent=self.root)
            return
        if not self.var_prod_name.get() or not self.txt_price.get() or not self.var_qty.get() or not self.txt_cost_price.get():
            # Removed barcode from required fields
            messagebox.showerror("Error", "All fields (except Barcode) are required", parent=self.root)
            return
        try:
            cost_price = self.get_cost_price_value()
            if cost_price is None:
                messagebox.showerror("Error", "Invalid cost price format", parent=self.root)
                return
            retail_price = self.get_price_value()
            if retail_price is None:
                messagebox.showerror("Error", "Invalid retail price format", parent=self.root)
                return
            qty = int(self.var_qty.get())
            margin = retail_price - cost_price
            status = self.determine_status(qty, self.var_status.get())
            self.var_status.set(status)
        except ValueError:
            messagebox.showerror("Error", "Price must be a number and Quantity must be an integer", parent=self.root)
            return
        
        con = sqlite3.connect(database=r'Possystem.db')
        cur = con.cursor()
        try:
            # Check if product name already exists
            cur.execute("SELECT * FROM product WHERE Name=?", (self.var_prod_name.get(),))
            row = cur.fetchone()
            if row:
                messagebox.showerror("Error", "Product already exists", parent=self.root)
                return
            
            # Only check barcode uniqueness if barcode is provided
            barcode = self.var_barcode.get().strip()
            if barcode:
                cur.execute("SELECT * FROM product WHERE barcode=?", (barcode,))
                row = cur.fetchone()
                if row:
                    messagebox.showerror("Error", "Barcode already exists for another product", parent=self.root)
                    return
            
            pid_value = self.var_pid.get().lstrip('0') or '0'
            if not pid_value.isdigit():
                pid_value = '0'
            self.ensure_database_columns()
            cur.execute("""
                INSERT INTO product (pid, Category, Name, barcode, cost_price, retailprice, margin, Quantity, Status) 
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (int(pid_value), self.var_ProdCat.get(), 
                  self.var_prod_name.get(), barcode, cost_price, retail_price, margin, qty, self.var_status.get()))
            con.commit()
            messagebox.showinfo("Success", "Product added successfully", parent=self.root)
            self.show_products()
            self.generate_pid()
            self.clear_product()
            self.update_low_stock_count()
        except Exception as ex:
            messagebox.showerror("Error", f"Error adding product: {str(ex)}", parent=self.root)
            print(f"Detailed error: {traceback.format_exc()}")
        finally:
            con.close()
    
    def update_product(self):
        if not self.var_pid.get():
            messagebox.showerror("Error", "Select a product to update", parent=self.root)
            return
        if self.var_ProdCat.get() == "" or self.var_ProdCat.get() == "Select":
            messagebox.showerror("Error", "Category is required", parent=self.root)
            return
        if not self.var_prod_name.get() or not self.txt_price.get() or not self.var_qty.get() or not self.txt_cost_price.get():
            # Removed barcode from required fields
            messagebox.showerror("Error", "All fields (except Barcode) are required", parent=self.root)
            return
        try:
            cost_price = self.get_cost_price_value()
            if cost_price is None:
                messagebox.showerror("Error", "Invalid cost price format", parent=self.root)
                return
            retail_price = self.get_price_value()
            if retail_price is None:
                messagebox.showerror("Error", "Invalid retail price format", parent=self.root)
                return
            qty = int(self.var_qty.get())
            margin = retail_price - cost_price
            status = self.determine_status(qty, self.var_status.get())
            self.var_status.set(status)
        except ValueError:
            messagebox.showerror("Error", "Price must be a number and Quantity must be an integer", parent=self.root)
            return
        
        con = sqlite3.connect(database=r'Possystem.db')
        cur = con.cursor()
        try:
            pid_value = self.var_pid.get().lstrip('0') or '0'
            if not pid_value.isdigit():
                messagebox.showerror("Error", "Invalid Product ID", parent=self.root)
                return
            pid_num = int(pid_value)
            
            # Check product name uniqueness excluding current product
            cur.execute("SELECT * FROM product WHERE Name=? AND pid!=?", 
                       (self.var_prod_name.get(), pid_num))
            row = cur.fetchone()
            if row:
                messagebox.showerror("Error", "Product name already exists for another product", parent=self.root)
                return
            
            # Check barcode uniqueness only if new barcode is provided and different from current
            barcode = self.var_barcode.get().strip()
            if barcode:
                cur.execute("SELECT * FROM product WHERE barcode=? AND pid!=?", (barcode, pid_num))
                row = cur.fetchone()
                if row:
                    messagebox.showerror("Error", "Barcode already exists for another product", parent=self.root)
                    return
            
            cur.execute("""
                UPDATE product SET Category=?, Name=?, barcode=?, cost_price=?, retailprice=?, margin=?, Quantity=?, Status=? 
                WHERE pid=?
            """, (self.var_ProdCat.get(), self.var_prod_name.get(), barcode,
                  cost_price, retail_price, margin, qty, self.var_status.get(), pid_num))
            con.commit()
            messagebox.showinfo("Success", "Product updated successfully", parent=self.root)
            self.show_products()
            self.update_low_stock_count()
        except Exception as ex:
            messagebox.showerror("Error", f"Error updating product: {str(ex)}", parent=self.root)
        finally:
            con.close()
    
    def delete_product(self):
        if not self.var_pid.get():
            messagebox.showerror("Error", "Select a product to delete", parent=self.root)
            return
        con = sqlite3.connect(database=r'Possystem.db')
        cur = con.cursor()
        try:
            pid_value = self.var_pid.get().lstrip('0') or '0'
            if not pid_value.isdigit():
                messagebox.showerror("Error", "Invalid Product ID", parent=self.root)
                return
            pid_num = int(pid_value)
            confirm = messagebox.askyesno("Confirm", "Do you really want to delete this product?", parent=self.root)
            if confirm:
                cur.execute("DELETE FROM product WHERE pid=?", (pid_num,))
                con.commit()
                messagebox.showinfo("Success", "Product deleted successfully", parent=self.root)
                self.clear_product()
                self.renumber_pids()
                self.generate_pid()
                self.show_products()
                self.update_low_stock_count()
        except Exception as ex:
            messagebox.showerror("Error", f"Error deleting product: {str(ex)}", parent=self.root)
        finally:
            con.close()
    
    def renumber_pids(self):
        try:
            con = sqlite3.connect(database=r'Possystem.db')
            cur = con.cursor()
            cur.execute("SELECT * FROM product ORDER BY pid")
            rows = cur.fetchall()
            for new_pid, row in enumerate(rows, 1):
                cur.execute("UPDATE product SET pid=? WHERE pid=?", (new_pid, row[0]))
            con.commit()
        except Exception as ex:
            messagebox.showerror("Error", f"Error renumbering products: {str(ex)}", parent=self.root)
        finally:
            con.close()
    
    def clear_product(self):
        self.var_prod_name.set("")
        self.var_barcode.set("")
        self.txt_price.delete(0, END)
        self.txt_cost_price.delete(0, END)
        self.var_margin.set("")
        self.var_qty.set("")
        self.var_ProdCat.set("")
        self.var_status.set("Active")
        self.var_searchtxt.set("")
        self.var_searchby.set("Select")
        self.Product_Table.selection_remove(self.Product_Table.selection())
        self.generate_pid()
        self.price_without_format = ""
        self.cost_price_without_format = ""
    
    def get_product_data(self, ev):
        f = self.Product_Table.focus()
        content = self.Product_Table.item(f)
        row = content['values']
        if row and len(row) >= 9:
            pid_value = row[0]
            self.var_pid.set(f"{int(pid_value):03d}")
            self.var_ProdCat.set(row[1])
            self.var_prod_name.set(row[2])
            self.var_barcode.set(row[3])   # barcode
            try:
                cost_str = str(row[4]).replace('Rs.', '').replace(',', '').strip()
                cost_value = float(cost_str)
                formatted_cost = f"{cost_value:,.2f}"
                self.txt_cost_price.delete(0, END)
                self.txt_cost_price.insert(0, formatted_cost.replace('.00', ''))
            except:
                self.txt_cost_price.delete(0, END)
                self.txt_cost_price.insert(0, row[4])
            try:
                price_str = str(row[5]).replace('Rs.', '').replace(',', '').strip()
                price_value = float(price_str)
                formatted_price = f"{price_value:,.2f}"
                self.txt_price.delete(0, END)
                self.txt_price.insert(0, formatted_price.replace('.00', ''))
            except:
                self.txt_price.delete(0, END)
                self.txt_price.insert(0, row[5])
            try:
                margin_str = str(row[6]).replace('Rs.', '').replace(',', '').strip()
                margin_value = float(margin_str)
                formatted_margin = f"{margin_value:,.2f}"
                self.var_margin.set(formatted_margin)
            except:
                self.var_margin.set(row[6])
            self.var_qty.set(row[7])
            self.var_status.set(row[8])
    
    def show_products(self):
        con = sqlite3.connect(database=r'Possystem.db')
        cur = con.cursor()
        try:
            self.ensure_database_columns()
            cur.execute("PRAGMA table_info(product)")
            columns = [col[1] for col in cur.fetchall()]
            pid_col = 'pid' if 'pid' in columns else 'rowid'
            cat_col = 'Category' if 'Category' in columns else 'category'
            name_col = 'Name' if 'Name' in columns else 'name'
            barcode_col = 'barcode' if 'barcode' in columns else None
            cost_col = 'cost_price' if 'cost_price' in columns else None
            retail_col = 'retailprice' if 'retailprice' in columns else None
            margin_col = 'margin' if 'margin' in columns else None
            qty_col = 'Quantity' if 'Quantity' in columns else 'quantity'
            status_col = 'Status' if 'Status' in columns else 'status'
            
            if barcode_col and cost_col and retail_col and margin_col:
                query = f"""
                    SELECT {pid_col}, {cat_col}, {name_col}, {barcode_col}, {cost_col}, {retail_col}, {margin_col}, {qty_col}, {status_col} 
                    FROM product 
                    ORDER BY {pid_col}
                """
                cur.execute(query)
            else:
                # Fallback if barcode column missing (shouldn't happen due to ensure_database_columns)
                query = "SELECT * FROM product ORDER BY pid"
                cur.execute(query)
            
            rows = cur.fetchall()
            self.Product_Table.delete(*self.Product_Table.get_children())
            
            for row in rows:
                try:
                    if len(row) >= 9:
                        try:
                            formatted_pid = f"{int(row[0]):03d}"
                        except (ValueError, TypeError):
                            formatted_pid = str(row[0])
                        category_val = row[1] if row[1] else ""
                        name_val = row[2] if row[2] else ""
                        barcode_val = row[3] if row[3] else ""   # can be empty
                        try:
                            cost_val = float(row[4]) if row[4] else 0
                            formatted_cost = f"Rs.{cost_val:,.2f}"
                        except (ValueError, TypeError):
                            formatted_cost = f"Rs.{row[4]}" if row[4] else "Rs.0.00"
                        try:
                            retail_val = float(row[5]) if row[5] else 0
                            formatted_retail = f"Rs.{retail_val:,.2f}"
                        except (ValueError, TypeError):
                            formatted_retail = f"Rs.{row[5]}" if row[5] else "Rs.0.00"
                        try:
                            margin_val = float(row[6]) if row[6] else 0
                            formatted_margin = f"Rs.{margin_val:,.2f}"
                        except (ValueError, TypeError):
                            cost_val = float(row[4]) if row[4] else 0
                            retail_val = float(row[5]) if row[5] else 0
                            margin_val = retail_val - cost_val
                            formatted_margin = f"Rs.{margin_val:,.2f}"
                        qty_val = int(row[7]) if row[7] else 0
                        status_val = row[8] if row[8] else "Active"
                        if qty_val <= self.reorder_threshold and status_val != "Inactive":
                            status_val = "Reorder"
                        tag = ''
                        if status_val == "Reorder":
                            tag = 'reorder'
                        elif status_val == "Active":
                            tag = 'active'
                        elif status_val == "Inactive":
                            tag = 'inactive'
                        formatted_row = (
                            formatted_pid,
                            str(category_val),
                            str(name_val),
                            str(barcode_val),
                            formatted_cost,
                            formatted_retail,
                            formatted_margin,
                            str(qty_val),
                            str(status_val)
                        )
                        self.Product_Table.insert('', END, values=formatted_row, tags=(tag,))
                    else:
                        # Handle old rows with fewer columns (if any)
                        continue
                except Exception as row_error:
                    print(f"Error processing row: {row_error}")
                    continue
            print(f"Successfully loaded {len(rows)} products")
        except Exception as ex:
            messagebox.showerror("Error", f"Error loading products: {str(ex)}", parent=self.root)
            print(f"Detailed error: {traceback.format_exc()}")
        finally:
            con.close()
    
    def show_products_by_category(self, category_name):
        con = sqlite3.connect(database=r'Possystem.db')
        cur = con.cursor()
        try:
            self.ensure_database_columns()
            cur.execute("PRAGMA table_info(product)")
            columns = [col[1] for col in cur.fetchall()]
            pid_col = 'pid' if 'pid' in columns else 'rowid'
            cat_col = 'Category' if 'Category' in columns else 'category'
            name_col = 'Name' if 'Name' in columns else 'name'
            barcode_col = 'barcode' if 'barcode' in columns else None
            cost_col = 'cost_price' if 'cost_price' in columns else None
            retail_col = 'retailprice' if 'retailprice' in columns else None
            margin_col = 'margin' if 'margin' in columns else None
            qty_col = 'Quantity' if 'Quantity' in columns else 'quantity'
            status_col = 'Status' if 'Status' in columns else 'status'
            
            if barcode_col and cost_col and retail_col and margin_col:
                query = f"""
                    SELECT {pid_col}, {cat_col}, {name_col}, {barcode_col}, {cost_col}, {retail_col}, {margin_col}, {qty_col}, {status_col} 
                    FROM product 
                    WHERE {cat_col}=? 
                    ORDER BY {pid_col}
                """
                cur.execute(query, (category_name,))
            else:
                query = "SELECT * FROM product WHERE Category=? ORDER BY pid"
                cur.execute(query, (category_name,))
            rows = cur.fetchall()
            self.Product_Table.delete(*self.Product_Table.get_children())
            for row in rows:
                if len(row) >= 9:
                    try:
                        try:
                            formatted_pid = f"{int(row[0]):03d}"
                        except:
                            formatted_pid = str(row[0])
                        category_val = row[1] if len(row) > 1 else ""
                        name_val = row[2] if len(row) > 2 else ""
                        barcode_val = row[3] if len(row) > 3 else ""
                        cost_val = 0
                        if len(row) > 4:
                            try:
                                cost_val = float(row[4])
                                formatted_cost = f"Rs.{cost_val:,.2f}"
                            except:
                                formatted_cost = f"Rs.{row[4]}"
                        else:
                            formatted_cost = "Rs.0.00"
                        retail_val = 0
                        if len(row) > 5:
                            try:
                                retail_val = float(row[5])
                                formatted_retail = f"Rs.{retail_val:,.2f}"
                            except:
                                formatted_retail = f"Rs.{row[5]}"
                        else:
                            formatted_retail = "Rs.0.00"
                        if len(row) > 6:
                            try:
                                margin_val = float(row[6])
                                formatted_margin = f"Rs.{margin_val:,.2f}"
                            except:
                                margin_val = retail_val - cost_val
                                formatted_margin = f"Rs.{margin_val:,.2f}"
                        else:
                            formatted_margin = "Rs.0.00"
                        qty_val = int(row[7]) if len(row) > 7 and row[7] else 0
                        status_val = row[8] if len(row) > 8 and row[8] else "Active"
                        if qty_val <= self.reorder_threshold and status_val != "Inactive":
                            status_val = "Reorder"
                        tag = ''
                        if status_val == "Reorder":
                            tag = 'reorder'
                        elif status_val == "Active":
                            tag = 'active'
                        elif status_val == "Inactive":
                            tag = 'inactive'
                        formatted_row = (
                            formatted_pid,
                            str(category_val),
                            str(name_val),
                            str(barcode_val),
                            formatted_cost,
                            formatted_retail,
                            formatted_margin,
                            str(qty_val),
                            str(status_val)
                        )
                        self.Product_Table.insert('', END, values=formatted_row, tags=(tag,))
                    except Exception as row_error:
                        print(f"Error processing row: {row_error}")
                        continue
        except Exception as ex:
            messagebox.showerror("Error", f"Error loading products: {str(ex)}", parent=self.root)
        finally:
            con.close()
    
    def search_product(self):
        if self.var_searchby.get() == "Select":
            messagebox.showerror("Error", "Select search criteria", parent=self.root)
            return
        if not self.var_searchtxt.get().strip() and self.var_searchby.get() != "Low Stock (≤5)":
            self.show_products()
            return
        con = sqlite3.connect(database=r'Possystem.db')
        cur = con.cursor()
        try:
            search_by = self.var_searchby.get()
            search_text = f"%{self.var_searchtxt.get().strip()}%"
            cur.execute("PRAGMA table_info(product)")
            columns = [col[1] for col in cur.fetchall()]
            pid_col = 'pid' if 'pid' in columns else 'rowid'
            cat_col = 'Category' if 'Category' in columns else 'category'
            name_col = 'Name' if 'Name' in columns else 'name'
            barcode_col = 'barcode' if 'barcode' in columns else None
            cost_col = 'cost_price' if 'cost_price' in columns else None
            retail_col = 'retailprice' if 'retailprice' in columns else None
            margin_col = 'margin' if 'margin' in columns else None
            qty_col = 'Quantity' if 'Quantity' in columns else 'quantity'
            status_col = 'Status' if 'Status' in columns else 'status'
            
            # We'll keep search limited to Category, Name, Status, Low Stock (Barcode not added to search options)
            if search_by == "Low Stock (≤5)":
                query = f"""
                    SELECT {pid_col}, {cat_col}, {name_col}, {barcode_col}, {cost_col}, {retail_col}, {margin_col}, {qty_col}, {status_col} 
                    FROM product 
                    WHERE {qty_col} <= ? AND ({status_col} != 'Inactive' OR {status_col} IS NULL)
                    ORDER BY {qty_col} ASC
                """
                cur.execute(query, (self.reorder_threshold,))
            elif search_by == "Status":
                query = f"""
                    SELECT {pid_col}, {cat_col}, {name_col}, {barcode_col}, {cost_col}, {retail_col}, {margin_col}, {qty_col}, {status_col} 
                    FROM product 
                    WHERE {status_col} LIKE ? 
                    ORDER BY {pid_col}
                """
                cur.execute(query, (search_text,))
            elif search_by == "Category":
                query = f"""
                    SELECT {pid_col}, {cat_col}, {name_col}, {barcode_col}, {cost_col}, {retail_col}, {margin_col}, {qty_col}, {status_col} 
                    FROM product 
                    WHERE {cat_col} LIKE ? 
                    ORDER BY {pid_col}
                """
                cur.execute(query, (search_text,))
            elif search_by == "Name":
                query = f"""
                    SELECT {pid_col}, {cat_col}, {name_col}, {barcode_col}, {cost_col}, {retail_col}, {margin_col}, {qty_col}, {status_col} 
                    FROM product 
                    WHERE {name_col} LIKE ? 
                    ORDER BY {pid_col}
                """
                cur.execute(query, (search_text,))
            else:
                # Fallback
                cur.execute("SELECT * FROM product WHERE Category LIKE ? OR Name LIKE ? OR Status LIKE ? ORDER BY pid", 
                           (search_text, search_text, search_text))
            
            rows = cur.fetchall()
            self.Product_Table.delete(*self.Product_Table.get_children())
            if rows:
                for row in rows:
                    if len(row) >= 9:
                        try:
                            try:
                                formatted_pid = f"{int(row[0]):03d}"
                            except:
                                formatted_pid = str(row[0])
                            category_val = row[1] if len(row) > 1 else ""
                            name_val = row[2] if len(row) > 2 else ""
                            barcode_val = row[3] if len(row) > 3 else ""
                            cost_val = 0
                            if len(row) > 4:
                                try:
                                    cost_val = float(row[4])
                                    formatted_cost = f"Rs.{cost_val:,.2f}"
                                except:
                                    formatted_cost = f"Rs.{row[4]}"
                            else:
                                formatted_cost = "Rs.0.00"
                            retail_val = 0
                            if len(row) > 5:
                                try:
                                    retail_val = float(row[5])
                                    formatted_retail = f"Rs.{retail_val:,.2f}"
                                except:
                                    formatted_retail = f"Rs.{row[5]}"
                            else:
                                formatted_retail = "Rs.0.00"
                            if len(row) > 6:
                                try:
                                    margin_val = float(row[6])
                                    formatted_margin = f"Rs.{margin_val:,.2f}"
                                except:
                                    margin_val = retail_val - cost_val
                                    formatted_margin = f"Rs.{margin_val:,.2f}"
                            else:
                                formatted_margin = "Rs.0.00"
                            qty_val = int(row[7]) if len(row) > 7 and row[7] else 0
                            status_val = row[8] if len(row) > 8 and row[8] else "Active"
                            if qty_val <= self.reorder_threshold and status_val != "Inactive":
                                status_val = "Reorder"
                            tag = ''
                            if status_val == "Reorder":
                                tag = 'reorder'
                            elif status_val == "Active":
                                tag = 'active'
                            elif status_val == "Inactive":
                                tag = 'inactive'
                            formatted_row = (
                                formatted_pid,
                                str(category_val),
                                str(name_val),
                                str(barcode_val),
                                formatted_cost,
                                formatted_retail,
                                formatted_margin,
                                str(qty_val),
                                str(status_val)
                            )
                            self.Product_Table.insert('', END, values=formatted_row, tags=(tag,))
                        except Exception as row_error:
                            print(f"Error processing row: {row_error}")
                            continue
            else:
                messagebox.showinfo("No Results", "No products found", parent=self.root)
        except Exception as ex:
            messagebox.showerror("Error", f"Error searching: {str(ex)}", parent=self.root)
        finally:
            con.close()
    
    # ========== PDF PRINTING ==========
    def print_all_products(self):
        try:
            con = sqlite3.connect(database=r'Possystem.db')
            cur = con.cursor()
            cur.execute("PRAGMA table_info(product)")
            columns = [col[1] for col in cur.fetchall()]
            pid_col = 'pid' if 'pid' in columns else 'rowid'
            cat_col = 'Category' if 'Category' in columns else 'category'
            name_col = 'Name' if 'Name' in columns else 'name'
            barcode_col = 'barcode' if 'barcode' in columns else None
            cost_col = 'cost_price' if 'cost_price' in columns else None
            retail_col = 'retailprice' if 'retailprice' in columns else None
            margin_col = 'margin' if 'margin' in columns else None
            qty_col = 'Quantity' if 'Quantity' in columns else 'quantity'
            status_col = 'Status' if 'Status' in columns else 'status'
            
            if barcode_col and cost_col and retail_col and margin_col:
                query = f"""
                    SELECT {pid_col}, {cat_col}, {name_col}, {barcode_col}, {cost_col}, {retail_col}, {margin_col}, {qty_col}, {status_col} 
                    FROM product 
                    ORDER BY {cat_col}, {name_col}
                """
                cur.execute(query)
            else:
                query = "SELECT * FROM product ORDER BY Category, Name"
                cur.execute(query)
            products = cur.fetchall()
            if not products:
                messagebox.showinfo("No Data", "No products to print", parent=self.root)
                return
            default_filename = f"Inventory_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            filename = filedialog.asksaveasfilename(
                parent=self.root,
                title="Save PDF Report",
                defaultextension=".pdf",
                initialfile=default_filename,
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
            )
            if not filename:
                return
            doc = SimpleDocTemplate(filename, pagesize=landscape(letter), 
                                   leftMargin=0.5*inch, rightMargin=0.5*inch,
                                   topMargin=0.5*inch, bottomMargin=0.8*inch)
            story = []
            styles = getSampleStyleSheet()
            footer_style = ParagraphStyle(
                name='Footer',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.HexColor('#666666'),
                alignment=TA_CENTER,
                spaceBefore=10,
                spaceAfter=10
            )
            disclaimer_style = ParagraphStyle(
                name='Disclaimer',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.HexColor('#856404'),
                backColor=colors.HexColor('#fff3cd'),
                alignment=TA_LEFT,
                spaceBefore=5,
                spaceAfter=5,
                borderWidth=1,
                borderColor=colors.HexColor('#ffeeba'),
                borderPadding=5
            )
            warning_style = ParagraphStyle(
                name='Warning',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#dc3545'),
                alignment=TA_CENTER,
                spaceBefore=10,
                spaceAfter=5,
                fontName='Helvetica-Bold'
            )
            summary_style = ParagraphStyle(
                name='Summary',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#9B59B6'),
                backColor=colors.HexColor('#F5E6FF'),
                alignment=TA_LEFT,
                spaceBefore=5,
                spaceAfter=5,
                borderWidth=1,
                borderColor=colors.HexColor('#9B59B6'),
                borderPadding=8,
                fontName='Helvetica-Bold'
            )
            title_style = ParagraphStyle(
                name='CustomTitle',
                parent=styles['Title'],
                fontSize=22,
                textColor=colors.HexColor('#9B59B6'),
                alignment=TA_CENTER,
                spaceAfter=10
            )

            title = Paragraph("INVENTORY PRODUCTS REPORT", title_style)
            story.append(title)
            story.append(Spacer(1, 10))
            date_str = datetime.now().strftime("%B %d, %Y at %I:%M %p")
            date_para = Paragraph(f"Generated on: {date_str}", styles['Normal'])
            story.append(date_para)
            story.append(Spacer(1, 5))
            company_info = Paragraph("Dukes Tech Services", styles['Heading3'])
            story.append(company_info)
            story.append(Spacer(1, 10))
            total_cost_value = 0
            total_retail_value = 0
            total_margin_value = 0
            total_quantity = 0
            reorder_count = 0
            for product in products:
                if len(product) >= 9:
                    try:
                        cost_price = float(product[4]) if len(product) > 4 and product[4] else 0
                        qty = int(product[7]) if len(product) > 7 and product[7] else 0
                        total_cost_value += cost_price * qty
                    except:
                        pass
                    try:
                        retail_price = float(product[5]) if len(product) > 5 and product[5] else 0
                        qty = int(product[7]) if len(product) > 7 and product[7] else 0
                        total_retail_value += retail_price * qty
                    except:
                        pass
                    try:
                        margin = float(product[6]) if len(product) > 6 and product[6] else 0
                        if margin == 0:
                            cost_price = float(product[4]) if len(product) > 4 and product[4] else 0
                            retail_price = float(product[5]) if len(product) > 5 and product[5] else 0
                            margin = retail_price - cost_price
                        qty = int(product[7]) if len(product) > 7 and product[7] else 0
                        total_margin_value += margin * qty
                    except:
                        pass
                    try:
                        qty = int(product[7]) if len(product) > 7 and product[7] else 0
                        total_quantity += qty
                    except:
                        pass
                    try:
                        qty = int(product[7]) if len(product) > 7 and product[7] else 0
                        status = product[8] if len(product) > 8 and product[8] else "Active"
                        if qty <= self.reorder_threshold and status != "Inactive":
                            reorder_count += 1
                    except:
                        pass
            summary_box = f"""
            <b>📊 SUMMARY REPORT</b><br/>
            <b>Total Products:</b> {len(products)}<br/>
            <b>Total Quantity:</b> {total_quantity}<br/>
            <b>Total Cost Value:</b> Rs. {total_cost_value:,.2f}<br/>
            <b>Total Retail Value:</b> Rs. {total_retail_value:,.2f}<br/>
            <b>Total Margin:</b> Rs. {total_margin_value:,.2f}<br/>
            <b>Profit Margin %:</b> {((total_retail_value - total_cost_value) / total_cost_value * 100) if total_cost_value > 0 else 0:,.2f}%<br/>
            <b><font color='#856404'>⚠️ Products Needing Reorder (Qty ≤ {self.reorder_threshold}): {reorder_count}</font></b>
            """
            summary = Paragraph(summary_box, summary_style)
            story.append(summary)
            story.append(Spacer(1, 15))
            disclaimer_text = f"""
            <font color="#856404"><b>⚠️ REORDER STATUS DISCLAIMER:</b></font><br/>
            Products with quantity of <b>{self.reorder_threshold} or less</b> are automatically flagged as 
            <font color="#856404"><b>'Reorder'</b></font> in the <b>Status column</b> of the table below. 
            These items require immediate attention to prevent stockouts. Please review low stock items and place orders as needed.
            """
            disclaimer = Paragraph(disclaimer_text, disclaimer_style)
            story.append(disclaimer)
            story.append(Spacer(1, 15))
            # Updated headers with Barcode
            data = [["ID", "Category", "Product Name", "Barcode", "Cost Price (Rs.)", "Retail Price (Rs.)", "Margin (Rs.)", "Qty", "Status"]]
            for product in products:
                if len(product) >= 9:
                    pid = str(product[0])
                    category = product[1] if len(product) > 1 else ""
                    name = product[2] if len(product) > 2 else ""
                    barcode = product[3] if len(product) > 3 else ""    # can be empty
                    try:
                        cost_price = float(product[4]) if len(product) > 4 and product[4] else 0
                        cost_formatted = f"{cost_price:,.2f}"
                    except:
                        cost_formatted = str(product[4]) if len(product) > 4 else "0.00"
                    try:
                        retail_price = float(product[5]) if len(product) > 5 and product[5] else 0
                        retail_formatted = f"{retail_price:,.2f}"
                    except:
                        retail_formatted = str(product[5]) if len(product) > 5 else "0.00"
                    try:
                        margin = float(product[6]) if len(product) > 6 and product[6] else 0
                        if margin == 0:
                            cost_price = float(product[4]) if len(product) > 4 and product[4] else 0
                            retail_price = float(product[5]) if len(product) > 5 and product[5] else 0
                            margin = retail_price - cost_price
                        margin_formatted = f"{margin:,.2f}"
                    except:
                        margin_formatted = "0.00"
                    try:
                        qty = int(product[7]) if len(product) > 7 and product[7] else 0
                        qty_str = str(qty)
                    except:
                        qty_str = str(product[7]) if len(product) > 7 else "0"
                    status = product[8] if len(product) > 8 and product[8] else "Active"
                    if qty <= self.reorder_threshold and status != "Inactive":
                        status = "Reorder"
                    data.append([pid, category, name, barcode, cost_formatted, retail_formatted, margin_formatted, qty_str, status])
            table = Table(data)
            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9B59B6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f5f7fa')),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (4, 1), (6, -1), 'RIGHT'),  # Cost, Retail, Margin
                ('ALIGN', (7, 1), (7, -1), 'RIGHT'),  # Qty
            ])
            for i, row in enumerate(data[1:], start=1):
                if row[8] == "Reorder":   # Status column index 8
                    style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#fff3cd'))
                    style.add('TEXTCOLOR', (0, i), (-1, i), colors.HexColor('#856404'))
            table.setStyle(style)
            story.append(table)
            if reorder_count > 0:
                story.append(Spacer(1, 10))
                warning_text = f"""
                <font color="#dc3545"><b>⚠️ URGENT: {reorder_count} product(s) require immediate reorder! Check the Status column above.</b></font>
                """
                warning = Paragraph(warning_text, warning_style)
                story.append(warning)
            story.append(Spacer(1, 20))
            footer_line1 = "Software Produced by Dukes Tech Services"
            footer_line2 = "Website: dukestechservices.com | Phone: +923097671363"
            footer_para1 = Paragraph(footer_line1, footer_style)
            footer_para2 = Paragraph(footer_line2, footer_style)
            story.append(footer_para1)
            story.append(footer_para2)
            def add_footer(canvas, doc):
                canvas.saveState()
                page_width, page_height = landscape(letter)
                footer_y = 0.4 * inch
                canvas.setFont('Helvetica', 8)
                canvas.setFillColor(colors.HexColor('#666666'))
                canvas.drawCentredString(page_width / 2, footer_y + 10, 
                                       "Software Produced by Dukes Tech Services")
                canvas.setFont('Helvetica', 8)
                canvas.setFillColor(colors.HexColor('#666666'))
                canvas.drawCentredString(page_width / 2, footer_y, 
                                       "Website: dukestechservices.com | Phone: +923097671363")
                canvas.setFont('Helvetica', 7)
                canvas.setFillColor(colors.HexColor('#856404'))
                canvas.drawCentredString(page_width / 2, footer_y - 15, 
                                       f"* Products with quantity ≤ {self.reorder_threshold} are flagged as 'Reorder' in the Status column")
                page_num = canvas.getPageNumber()
                canvas.setFont('Helvetica', 8)
                canvas.setFillColor(colors.HexColor('#666666'))
                canvas.drawRightString(page_width - 0.5 * inch, footer_y + 10, 
                                     f"Page {page_num}")
                canvas.restoreState()
            doc.build(story, onLaterPages=add_footer, onFirstPage=add_footer)
            open_pdf = messagebox.askyesno("PDF Generated", 
                                          f"PDF saved successfully at:\n{filename}\n\nDo you want to open it now?", 
                                          parent=self.root)
            if open_pdf:
                webbrowser.open(filename)
        except Exception as ex:
            messagebox.showerror("Error", f"Error generating PDF: {str(ex)}", parent=self.root)
            print(f"PDF Error: {traceback.format_exc()}")
        finally:
            con.close()

if __name__ == "__main__":
    root = Tk()
    obj = IntegratedInventorySystem(root)
    root.mainloop()